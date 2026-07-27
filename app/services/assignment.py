from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from decimal import Decimal, InvalidOperation
from io import BytesIO
import re

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import Batch, BatchItem, BatchStatus, BatchType, Product, ScanLog, Serial, SerialStatus, TransactionType, User, WarehouseLevel
from app.services.change_audit import record_change
from app.services.expiry import parse_optional_date
from app.services.inventory import InventoryError, create_batch, generate_serials, log_inventory_transaction, normalize_serial


MAX_ASSIGNMENT_QUANTITY = 5000
TALLY_LEDGER_ROW_RE = re.compile(r"\b(?P<tax>CGST|SGST|IGST|GST)\b\s*@\s*(?P<rate>\d+(?:\.\d+)?)\s*%", re.IGNORECASE)

PRODUCT_CODE_HEADERS = {"product code", "code", "product"}
PRODUCT_NAME_HEADERS = {
    "product name",
    "item name",
    "stock item",
    "tally stock item",
    "description of goods",
    "description",
    "goods",
}
QUANTITY_HEADERS = {"quantity", "qty", "billed quantity"}
RATE_HEADERS = {"rate", "unit rate", "price", "item rate"}
BATCH_HEADERS = {"batch", "batch no", "batch number", "product batch"}
MFG_DATE_HEADERS = {"mfg date", "manufacturing date", "manufacture date"}
EXPIRY_DATE_HEADERS = {"expiry date", "expiry", "exp date"}
WAREHOUSE_HEADERS = {"warehouse", "wh", "location"}
WAREHOUSE_LEVEL_HEADERS = {"warehouse level", "franchise level", "location level"}
HSN_HEADERS = {"hsn", "hsn sac", "hsn code"}
GST_HEADERS = {"gst", "gst rate", "gst percentage", "gst percent"}
CGST_HEADERS = {"cgst", "cgst rate", "cgst percentage", "cgst percent"}
SGST_HEADERS = {"sgst", "sgst rate", "sgst utgst rate", "sgst percentage", "sgst percent"}
IGST_HEADERS = {"igst", "ignst", "igst rate", "ignst rate", "igst percentage", "igst percent"}
UNIT_HEADERS = {"unit", "per", "uom", "item rate per"}


@dataclass(frozen=True)
class AssignmentLine:
    product: Product
    quantity: int
    rate: Decimal | None = None
    prefix: str | None = None
    product_batch_number: str | None = None
    mfg_date: date | None = None
    expiry_date: date | None = None
    warehouse: str | None = None
    warehouse_level: str = WarehouseLevel.COMPANY_WAREHOUSE.value


@dataclass(frozen=True)
class _AssignmentImportColumns:
    product_code_col: int | None
    product_name_col: int | None
    qty_col: int
    rate_col: int | None = None
    batch_col: int | None = None
    mfg_col: int | None = None
    expiry_col: int | None = None
    warehouse_col: int | None = None
    warehouse_level_col: int | None = None
    hsn_col: int | None = None
    gst_col: int | None = None
    cgst_col: int | None = None
    sgst_col: int | None = None
    igst_col: int | None = None
    unit_col: int | None = None


@dataclass(frozen=True)
class _RawAssignmentLine:
    row_number: int
    product_code: str | None
    product_name: str | None
    quantity: int
    rate: Decimal | None = None
    product_batch_number: str | None = None
    mfg_date: date | None = None
    expiry_date: date | None = None
    warehouse: str | None = None
    warehouse_level: str = WarehouseLevel.COMPANY_WAREHOUSE.value
    hsn: str | None = None
    gst_rate: Decimal | None = None
    cgst_rate: Decimal | None = None
    sgst_rate: Decimal | None = None
    igst_rate: Decimal | None = None
    unit: str | None = None


def parse_bulk_assignment_xlsx(
    db: Session,
    data: bytes,
    user: User | None = None,
    *,
    allow_product_create: bool = True,
) -> list[AssignmentLine]:
    try:
        workbook = load_workbook(BytesIO(data), data_only=True)
    except Exception as exc:
        raise InventoryError("Upload a readable Excel .xlsx file") from exc

    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise InventoryError("Excel file is empty")

    raw_lines = _parse_assignment_rows(rows)
    if not raw_lines:
        raise InventoryError("Excel file has no assignment rows")

    total = sum(line.quantity for line in raw_lines)
    if total > MAX_ASSIGNMENT_QUANTITY:
        raise InventoryError(f"Assign {MAX_ASSIGNMENT_QUANTITY} barcodes or fewer at a time")

    lines: list[AssignmentLine] = []
    for raw_line in raw_lines:
        product = _resolve_import_product(db, raw_line, user, allow_product_create)
        lines.append(
            AssignmentLine(
                product=product,
                quantity=raw_line.quantity,
                rate=raw_line.rate,
                product_batch_number=raw_line.product_batch_number,
                mfg_date=raw_line.mfg_date,
                expiry_date=raw_line.expiry_date,
                warehouse=raw_line.warehouse,
                warehouse_level=raw_line.warehouse_level,
            )
        )
    return lines


def assign_barcodes_to_existing_stock(
    db: Session,
    user: User,
    lines: list[AssignmentLine],
    notes: str | None = None,
    source: str = "MANUAL",
    initial_status: SerialStatus = SerialStatus.IN_STOCK,
) -> Batch:
    total = sum(line.quantity for line in lines)
    if total < 1:
        raise InventoryError("Quantity must be at least 1")
    if total > MAX_ASSIGNMENT_QUANTITY:
        raise InventoryError(f"Assign {MAX_ASSIGNMENT_QUANTITY} barcodes or fewer at a time")

    try:
        batch = create_batch(
            db,
            user,
            BatchType.QR_ASSIGNMENT,
            party_name="Existing Tally stock",
            notes=notes,
            reason_code=source,
            commit=False,
        )
        created_serials: list[Serial] = []
        for line in lines:
            created_serials.extend(
                generate_serials(
                    db,
                    line.product,
                    line.quantity,
                    prefix=line.prefix or line.product.product_code,
                    initial_status=initial_status,
                    product_batch_number=line.product_batch_number,
                    mfg_date=line.mfg_date,
                    expiry_date=line.expiry_date,
                    warehouse=line.warehouse,
                    warehouse_level=line.warehouse_level,
                    commit=False,
                )
            )

        message = (
            "Barcode assigned to existing Tally stock"
            if initial_status == SerialStatus.IN_STOCK
            else "Barcode generated for future stock movement"
        )
        for serial in created_serials:
            db.add(BatchItem(batch_id=batch.id, serial_id=serial.id))
            db.add(
                ScanLog(
                    serial_id=serial.id,
                    serial_number_raw=serial.serial_number,
                    user_id=user.id,
                    action=TransactionType.QR_ASSIGNMENT.value,
                    batch_id=batch.id,
                    status=initial_status.value,
                    message=message,
                )
            )
            log_inventory_transaction(
                db,
                user,
                TransactionType.QR_ASSIGNMENT,
                serial=serial,
                product=serial.product,
                batch=batch,
                status_from=None,
                status_to=initial_status.value,
                reason_code=source,
                notes=notes or message,
            )
        batch.status = BatchStatus.CLOSED.value
        batch.submitted_at = batch.created_at
        batch.synced_at = batch.created_at
        db.commit()
    except Exception:
        db.rollback()
        raise
    db.refresh(batch)
    return batch


def _parse_assignment_rows(rows: list[tuple[object, ...]]) -> list[_RawAssignmentLine]:
    header_match = _find_assignment_header(rows)
    if header_match:
        header_index, columns = header_match
        start_index = header_index + 1
    else:
        columns = _AssignmentImportColumns(product_code_col=0, product_name_col=None, qty_col=1)
        start_index = 0
    sheet_rates = _extract_sheet_tax_rates(rows)

    lines: list[_RawAssignmentLine] = []
    for row_index, row in enumerate(rows[start_index:], start=start_index + 1):
        product_code = _cell_text(row, columns.product_code_col)
        product_name = _cell_text(row, columns.product_name_col)
        raw_qty = _cell_value(row, columns.qty_col)
        if _is_blank_row(product_code, product_name, raw_qty):
            continue
        if _is_summary_or_ledger_row(product_code, product_name):
            continue
        if not product_code and not product_name:
            raise InventoryError(f"Row {row_index}: product name is required")
        quantity = _parse_quantity(raw_qty, row_index)

        try:
            mfg_date = parse_optional_date(_cell_value(row, columns.mfg_col)) if columns.mfg_col is not None else None
            expiry_date = (
                parse_optional_date(_cell_value(row, columns.expiry_col)) if columns.expiry_col is not None else None
            )
        except (TypeError, ValueError) as exc:
            raise InventoryError(f"Row {row_index}: use a valid date for mfg/expiry") from exc
        if mfg_date and expiry_date and expiry_date <= mfg_date:
            raise InventoryError(f"Row {row_index}: expiry date must be after mfg date")

        warehouse_level = _cell_text(row, columns.warehouse_level_col) or WarehouseLevel.COMPANY_WAREHOUSE.value
        try:
            warehouse_level = WarehouseLevel(warehouse_level).value
        except ValueError as exc:
            raise InventoryError(f"Row {row_index}: warehouse level is not recognized") from exc

        cgst_rate = _parse_decimal(_cell_value(row, columns.cgst_col)) or sheet_rates.get("cgst")
        sgst_rate = _parse_decimal(_cell_value(row, columns.sgst_col)) or sheet_rates.get("sgst")
        igst_rate = _parse_decimal(_cell_value(row, columns.igst_col)) or sheet_rates.get("igst")
        gst_rate = _parse_decimal(_cell_value(row, columns.gst_col)) or sheet_rates.get("gst")
        if gst_rate is None:
            if igst_rate is not None:
                gst_rate = igst_rate
            elif cgst_rate is not None or sgst_rate is not None:
                gst_rate = (cgst_rate or Decimal("0")) + (sgst_rate or Decimal("0"))

        lines.append(
            _RawAssignmentLine(
                row_number=row_index,
                product_code=normalize_serial(product_code) if product_code else None,
                product_name=product_name or None,
                quantity=quantity,
                rate=_parse_decimal(_cell_value(row, columns.rate_col)),
                product_batch_number=_cell_text(row, columns.batch_col) or None,
                mfg_date=mfg_date,
                expiry_date=expiry_date,
                warehouse=_cell_text(row, columns.warehouse_col) or None,
                warehouse_level=warehouse_level,
                hsn=_cell_text(row, columns.hsn_col) or None,
                gst_rate=gst_rate,
                cgst_rate=cgst_rate,
                sgst_rate=sgst_rate,
                igst_rate=igst_rate,
                unit=_cell_text(row, columns.unit_col) or None,
            )
        )
    return lines


def _find_assignment_header(rows: list[tuple[object, ...]]) -> tuple[int, _AssignmentImportColumns] | None:
    for row_index, row in enumerate(rows):
        header = [_normalize_header(value) for value in row]
        product_code_col = _find_column(header, PRODUCT_CODE_HEADERS)
        product_name_col = _find_column(header, PRODUCT_NAME_HEADERS)
        qty_col = _find_column(header, QUANTITY_HEADERS)
        if qty_col is None or (product_code_col is None and product_name_col is None):
            continue
        return (
            row_index,
            _AssignmentImportColumns(
                product_code_col=product_code_col,
                product_name_col=product_name_col,
                qty_col=qty_col,
                rate_col=_find_column(header, RATE_HEADERS),
                batch_col=_find_column(header, BATCH_HEADERS),
                mfg_col=_find_column(header, MFG_DATE_HEADERS),
                expiry_col=_find_column(header, EXPIRY_DATE_HEADERS),
                warehouse_col=_find_column(header, WAREHOUSE_HEADERS),
                warehouse_level_col=_find_column(header, WAREHOUSE_LEVEL_HEADERS),
                hsn_col=_find_column(header, HSN_HEADERS),
                gst_col=_find_column(header, GST_HEADERS),
                cgst_col=_find_column(header, CGST_HEADERS),
                sgst_col=_find_column(header, SGST_HEADERS),
                igst_col=_find_column(header, IGST_HEADERS),
                unit_col=_find_column(header, UNIT_HEADERS),
            ),
        )
    return None


def _resolve_import_product(
    db: Session,
    line: _RawAssignmentLine,
    user: User | None,
    allow_product_create: bool,
) -> Product:
    if line.product_code:
        product = db.scalar(select(Product).where(Product.product_code == line.product_code, Product.active.is_(True)))
        if not product:
            raise InventoryError(f"Row {line.row_number}: product {line.product_code} was not found")
        return product

    product_name = (line.product_name or "").strip()
    product = _find_active_product_by_name(db, product_name)
    if product:
        return product
    if not allow_product_create:
        raise InventoryError(f"Row {line.row_number}: product {product_name} was not found")

    product = Product(
        product_code=_next_import_product_code(db, product_name),
        product_name=product_name,
        hsn=line.hsn or "",
        gst_rate=float(line.gst_rate or Decimal("0")),
        unit=line.unit or "Pcs",
        default_rate=0,
        tally_stock_item_name=product_name,
    )
    db.add(product)
    db.flush()
    record_change(
        db,
        user,
        entity_type="product",
        entity_id=product.id,
        action="create",
        before=None,
        after=_product_snapshot(product),
    )
    return product


def _find_active_product_by_name(db: Session, product_name: str) -> Product | None:
    key = _normalize_lookup(product_name)
    if not key:
        return None
    rows = db.scalars(select(Product).where(Product.active.is_(True))).all()
    for product in rows:
        if key in {
            _normalize_lookup(product.product_name),
            _normalize_lookup(product.nickname),
            _normalize_lookup(product.tally_stock_item_name),
            _normalize_lookup(product.alternate_tally_stock_item_name),
        }:
            return product
    return None


def _next_import_product_code(db: Session, product_name: str) -> str:
    base = re.sub(r"[^A-Z0-9]+", "", product_name.upper())[:32] or "TALLYITEM"
    candidate = base
    suffix = 2
    while db.scalar(select(Product.id).where(Product.product_code == candidate)) is not None:
        suffix_text = f"-{suffix}"
        candidate = f"{base[: 80 - len(suffix_text)]}{suffix_text}"
        suffix += 1
    return candidate


def _extract_sheet_tax_rates(rows: list[tuple[object, ...]]) -> dict[str, Decimal]:
    rates: dict[str, Decimal] = {}
    for row in rows:
        for value in row:
            if not isinstance(value, str):
                continue
            match = TALLY_LEDGER_ROW_RE.search(value)
            if not match:
                continue
            tax_key = match.group("tax").lower()
            rates[tax_key] = Decimal(match.group("rate"))
    if "gst" not in rates:
        if "igst" in rates:
            rates["gst"] = rates["igst"]
        elif "cgst" in rates or "sgst" in rates:
            rates["gst"] = rates.get("cgst", Decimal("0")) + rates.get("sgst", Decimal("0"))
    return rates


def _parse_quantity(raw_qty: object, row_index: int) -> int:
    if raw_qty is None or raw_qty == "":
        raise InventoryError(f"Row {row_index}: quantity is required")
    if isinstance(raw_qty, str):
        match = re.search(r"-?\d+(?:\.\d+)?", raw_qty.replace(",", ""))
        if not match:
            raise InventoryError(f"Row {row_index}: quantity must be a whole number")
        value = Decimal(match.group(0))
    else:
        try:
            value = Decimal(str(raw_qty))
        except (InvalidOperation, ValueError) as exc:
            raise InventoryError(f"Row {row_index}: quantity must be a whole number") from exc
    if value != value.to_integral_value():
        raise InventoryError(f"Row {row_index}: quantity must be a whole number")
    quantity = int(value)
    if quantity < 1:
        raise InventoryError(f"Row {row_index}: quantity must be at least 1")
    return quantity


def _parse_decimal(value: object) -> Decimal | None:
    if value is None or value == "":
        return None
    if isinstance(value, str):
        match = re.search(r"-?\d+(?:\.\d+)?", value.replace(",", ""))
        if not match:
            return None
        value = match.group(0)
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None


def _cell_value(row: tuple[object, ...], index: int | None) -> object | None:
    if index is None or index >= len(row):
        return None
    return row[index]


def _cell_text(row: tuple[object, ...], index: int | None) -> str:
    value = _cell_value(row, index)
    return str(value or "").strip()


def _is_blank_row(product_code: str, product_name: str, raw_qty: object | None) -> bool:
    return not product_code and not product_name and raw_qty in {None, ""}


def _is_summary_or_ledger_row(product_code: str, product_name: str) -> bool:
    value = _normalize_lookup(product_name or product_code)
    compact_value = value.strip()
    if compact_value in {"total", "round off", "amount chargeable in words"}:
        return True
    return any(token in value for token in {" output cgst ", " output sgst ", " output igst "})


def _normalize_lookup(value: object | None) -> str:
    collapsed = re.sub(r"\s+", " ", str(value or "").strip()).casefold()
    return f" {collapsed} "


def _normalize_header(value: object | None) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").strip().lower()).strip()


def _product_snapshot(product: Product) -> dict[str, object]:
    return {
        "id": product.id,
        "product_code": product.product_code,
        "product_name": product.product_name,
        "nickname": product.nickname,
        "category": product.category,
        "brand": product.brand,
        "hsn": product.hsn,
        "gst_rate": product.gst_rate,
        "unit": product.unit,
        "default_rate": product.default_rate,
        "sales_discount_rate": product.sales_discount_rate,
        "shelf_verification_interval": product.shelf_verification_interval,
        "purchase_qr_print_allowed": product.purchase_qr_print_allowed,
        "tally_stock_item_name": product.tally_stock_item_name,
        "alternate_tally_stock_item_name": product.alternate_tally_stock_item_name,
        "active": product.active,
    }


def _find_column(header: list[str], names: set[str]) -> int | None:
    normalized_names = {_normalize_header(name) for name in names}
    for index, value in enumerate(header):
        if value in normalized_names:
            return index
    return None
