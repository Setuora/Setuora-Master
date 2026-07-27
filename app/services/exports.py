from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from xml.sax.saxutils import escape

import qrcode
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.platypus import Image, PageBreak, SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from PIL import Image as PILImage

from app.models import AuditFinding, Batch, InventoryTransaction, ScanLog, Serial, StorageLocation
from app.services.log_fields import barcode_sold_by, invoice_created_by, product_audited_by
from app.services.report_format import report_date

LABEL_WIDTH_MM = 48.5
LABEL_HEIGHT_MM = 25.4
QR_SIZE_MM = 19.5
DEFAULT_LABEL_ROWS = 11
DEFAULT_LABEL_COLUMNS = 4
MIN_LABEL_ROWS = 1
MAX_LABEL_ROWS = 11
MIN_LABEL_COLUMNS = 1
MAX_LABEL_COLUMNS = 4
DANGEROUS_SPREADSHEET_PREFIXES = ("=", "+", "-", "@", "\t", "\r")
SCAN_EXPORT_HEADERS = ["Date", "User", "Action", "Serial", "Status", "Batch", "Message", "Tally Reference"]
TRANSACTION_EXPORT_HEADERS = [
    "Date", "User", "Type", "Serial", "Product Code", "Product Name",
    "Invoice Created By", "Barcode Sold By", "Product Audited By",
    "From Status", "To Status", "Reason", "Batch/Reference",
    "Tally Reference", "Notes",
]
MISSING_STOCK_EXPORT_HEADERS = [
    "Audit Date", "Audited By", "Audit Batch", "Serial", "Product Code",
    "Product Name", "Product Batch", "Warehouse", "Storage Location",
    "Mfg Date", "Expiry Date", "Expected Status",
]
SERIAL_EXPORT_HEADERS = [
    "Product Code", "Product Name", "Tally Stock Item", "Serial Number",
    "Product Batch", "Mfg Date", "Expiry Date", "Warehouse",
    "Warehouse Level", "Status", "Created At",
]


def spreadsheet_safe(value):
    if value is None:
        return ""
    if isinstance(value, (int, float, date, datetime)):
        return value
    text = str(value)
    if text.startswith(DANGEROUS_SPREADSHEET_PREFIXES):
        return f"'{text}"
    return text


def safe_row(values):
    return [spreadsheet_safe(value) for value in values]


def select_export_columns(
    headers: list[str],
    rows: list[list[object]],
    fields: list[str] | None = None,
) -> tuple[list[str], list[list[object]]]:
    requested = {field.strip() for field in fields or [] if field.strip()}
    if not requested:
        return headers, rows
    indexes = [index for index, header in enumerate(headers) if header.strip() in requested]
    if not indexes:
        return headers, rows
    return (
        [headers[index] for index in indexes],
        [[row[index] if index < len(row) else "" for index in indexes] for row in rows],
    )


def barcode_png(value: str) -> bytes:
    qr = qrcode.QRCode(
        version=None,
        error_correction=qrcode.constants.ERROR_CORRECT_Q,
        box_size=12,
        border=4,
    )
    qr.add_data(value)
    qr.make(fit=True)
    image = qr.make_image(fill_color="black", back_color="white").convert("RGB")
    stream = BytesIO()
    image.save(stream, format="PNG")
    return stream.getvalue()


def scans_xlsx(scans: list[ScanLog], fields: list[str] | None = None) -> bytes:
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Scans"
    rows = [
        safe_row(
            [
                report_date(scan.created_at),
                scan.user.username,
                scan.action,
                scan.serial_number_raw,
                scan.status,
                scan.batch.batch_number if scan.batch else "",
                scan.message or "",
                scan.tally_reference or "",
            ]
        )
        for scan in scans
    ]
    headers, rows = select_export_columns(SCAN_EXPORT_HEADERS, rows, fields)
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    for column in sheet.columns:
        width = max(len(str(cell.value or "")) for cell in column)
        sheet.column_dimensions[column[0].column_letter].width = min(max(width + 2, 12), 40)
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def serials_xlsx(serials: list[Serial], fields: list[str] | None = None) -> bytes:
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Barcodes"
    rows = []
    for serial in serials:
        product = serial.product
        rows.append(
            safe_row(
                [
                    product.product_code,
                    product.product_name,
                    product.tally_stock_item_name,
                    serial.serial_number,
                    serial.product_batch_number or "",
                    report_date(serial.mfg_date),
                    report_date(serial.expiry_date),
                    serial.warehouse or "",
                    serial.warehouse_level,
                    serial.display_status,
                    report_date(serial.created_at),
                ]
            )
        )
    headers, rows = select_export_columns(SERIAL_EXPORT_HEADERS, rows, fields)
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    _autosize(sheet)
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def transactions_xlsx(transactions: list[InventoryTransaction], fields: list[str] | None = None) -> bytes:
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Transactions"
    rows = [
        safe_row(
            [
                report_date(txn.created_at),
                txn.user.username,
                txn.transaction_type,
                txn.serial_number or "",
                txn.product.product_code if txn.product else "",
                txn.product.product_name if txn.product else "",
                invoice_created_by(txn),
                barcode_sold_by(txn),
                product_audited_by(txn),
                txn.status_from or "",
                txn.status_to or "",
                txn.reason_code or "",
                txn.reference_number or "",
                txn.tally_reference or "",
                txn.notes or "",
            ]
        )
        for txn in transactions
    ]
    headers, rows = select_export_columns(TRANSACTION_EXPORT_HEADERS, rows, fields)
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    _autosize(sheet)
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def missing_stock_xlsx(findings: list[AuditFinding], fields: list[str] | None = None) -> bytes:
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Missing Stock"
    rows = []
    for finding in findings:
        serial = finding.serial
        rows.append(
            safe_row(
                [
                    report_date(finding.created_at),
                    finding.batch.user.username,
                    finding.batch.batch_number,
                    finding.serial_number,
                    finding.product_code or "",
                    finding.product_name or "",
                    serial.product_batch_number if serial else "",
                    serial.warehouse if serial else "",
                    serial.location.full_path if serial and serial.location else "",
                    report_date(serial.mfg_date) if serial else "",
                    report_date(serial.expiry_date) if serial else "",
                    finding.expected_status or "",
                ]
            )
        )
    headers, rows = select_export_columns(MISSING_STOCK_EXPORT_HEADERS, rows, fields)
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    _autosize(sheet)
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def audit_reconciliation_xlsx(
    report: dict[str, object],
    fields: list[str] | None = None,
) -> bytes:
    from openpyxl import Workbook

    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    summary.append(["Period start", _datetime_text(report.get("start_at")) or "All"])
    summary.append(["Period end before", _datetime_text(report.get("end_at")) or "All"])
    summary.append(["Audit batches", report.get("audit_batch_count", 0)])
    summary.append(["Verified", report.get("verified", 0)])
    summary.append(["Missing", report.get("missing", 0)])
    summary.append(["Extra", report.get("extra", 0)])
    summary.append(["Total findings", report.get("total", 0)])
    summary.append(["Pending", report.get("pending", 0)])

    batches = workbook.create_sheet("Audit Batches")
    batches.append(["Audit Date", "Audit Batch", "Audited By", "Products", "Verified", "Missing", "Extra", "Total", "Pending"])
    for row in report.get("batch_rows", []):
        batches.append(
            safe_row(
                [
                    _datetime_text(row.get("audit_at")),
                    row.get("batch_number", ""),
                    row.get("audited_by", ""),
                    row.get("products", 0),
                    row.get("verified", 0),
                    row.get("missing", 0),
                    row.get("extra", 0),
                    row.get("total", 0),
                    row.get("pending", 0),
                ]
            )
        )

    products = workbook.create_sheet("Product Reconciliation")
    products.append(["Product Code", "Product Name", "Audit Batches", "Verified", "Missing", "Extra", "Total", "Pending"])
    for row in report.get("product_rows", []):
        products.append(
            safe_row(
                [
                    row.get("product_code", ""),
                    row.get("product_name", ""),
                    row.get("audit_batches", ""),
                    row.get("verified", 0),
                    row.get("missing", 0),
                    row.get("extra", 0),
                    row.get("total", 0),
                    row.get("pending", 0),
                ]
            )
        )

    findings = workbook.create_sheet("Serial Findings")
    findings.append(
        [
            "Audit Date",
            "Audit Batch",
            "Audited By",
            "Finding",
            "Serial",
            "Product Code",
            "Product Name",
            "Expected Status",
            "Scanned Status",
        ]
    )
    for row in report.get("finding_rows", []):
        findings.append(
            safe_row(
                [
                    _datetime_text(row.get("audit_at")),
                    row.get("batch_number", ""),
                    row.get("audited_by", ""),
                    row.get("type", ""),
                    row.get("serial_number", ""),
                    row.get("product_code", ""),
                    row.get("product_name", ""),
                    row.get("expected_status", ""),
                    row.get("scanned_status", ""),
                ]
            )
        )

    requested = {field.strip() for field in fields or [] if field.strip()}
    if requested:
        for sheet in list(workbook.worksheets[1:]):
            indexes = [
                cell.column
                for cell in sheet[1]
                if str(cell.value or "").strip() in requested
            ]
            if not indexes:
                workbook.remove(sheet)
                continue
            for column in range(sheet.max_column, 0, -1):
                if column not in indexes:
                    sheet.delete_cols(column)

    for sheet in workbook.worksheets:
        _autosize(sheet)
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _datetime_text(value: object) -> str:
    return report_date(value)


def _autosize(sheet) -> None:
    for column in sheet.columns:
        width = max(len(str(cell.value or "")) for cell in column)
        sheet.column_dimensions[column[0].column_letter].width = min(max(width + 2, 12), 42)


def label_layout(rows_per_page: int = DEFAULT_LABEL_ROWS, columns_per_page: int = DEFAULT_LABEL_COLUMNS) -> tuple[int, int]:
    rows = rows_per_page if MIN_LABEL_ROWS <= rows_per_page <= MAX_LABEL_ROWS else DEFAULT_LABEL_ROWS
    columns = columns_per_page if MIN_LABEL_COLUMNS <= columns_per_page <= MAX_LABEL_COLUMNS else DEFAULT_LABEL_COLUMNS
    return rows, columns


def _label_image(value: str, target_width: float, target_height: float | None = None) -> Image:
    png = barcode_png(value)
    px_w, px_h = PILImage.open(BytesIO(png)).size
    scale = target_width / px_w
    if target_height is not None:
        scale = min(scale, target_height / px_h)
    width = px_w * scale
    height = px_h * scale
    return Image(BytesIO(png), width=width, height=height)


def _label_cell(value: str, target_width: float, target_height: float) -> Table:
    qr_size = min(QR_SIZE_MM * mm, target_height, target_width * 0.48)
    text_width = max(10 * mm, target_width - qr_size - 2 * mm)
    qr_image = _label_image(value, qr_size, qr_size)
    serial_style = ParagraphStyle(
        "LabelSerial",
        fontName="Helvetica-Bold",
        fontSize=7.2,
        leading=8.2,
        textColor=colors.black,
        alignment=0,
    )
    label = Table(
        [[qr_image, Paragraph(escape(value), serial_style)]],
        colWidths=[qr_size, text_width],
        rowHeights=[target_height],
    )
    label.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (0, 0), (0, 0), "CENTER"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ]
        )
    )
    return label


def barcode_labels_pdf(
    serials: list[Serial],
    rows_per_page: int = DEFAULT_LABEL_ROWS,
    columns_per_page: int = DEFAULT_LABEL_COLUMNS,
) -> bytes:
    rows_per_page, columns_per_page = label_layout(rows_per_page, columns_per_page)
    stream = BytesIO()
    doc = SimpleDocTemplate(stream, pagesize=A4, rightMargin=8 * mm, leftMargin=8 * mm, topMargin=8 * mm, bottomMargin=8 * mm)
    story = []
    labels_per_page = rows_per_page * columns_per_page
    col_width = LABEL_WIDTH_MM * mm
    row_height = LABEL_HEIGHT_MM * mm
    cell_width = col_width - 2 * mm
    cell_height = row_height - 2 * mm

    for page_start in range(0, len(serials), labels_per_page):
        page_serials = serials[page_start:page_start + labels_per_page]
        table_rows = []
        for row_index in range(rows_per_page):
            row = []
            for column_index in range(columns_per_page):
                serial_index = row_index * columns_per_page + column_index
                if serial_index < len(page_serials):
                    row.append(_label_cell(page_serials[serial_index].serial_number, cell_width, cell_height))
                else:
                    row.append("")
            table_rows.append(row)
        table = Table(
            table_rows,
            colWidths=[col_width] * columns_per_page,
            rowHeights=[row_height] * rows_per_page,
        )
        table.setStyle(
            TableStyle(
                [
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 1 * mm),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 1 * mm),
                    ("TOPPADDING", (0, 0), (-1, -1), 1 * mm),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 1 * mm),
                ]
            )
        )
        story.append(table)
        if page_start + labels_per_page < len(serials):
            story.append(PageBreak())
    if not story:
        styles = getSampleStyleSheet()
        story = [Paragraph("No labels selected", styles["BodyText"])]
    doc.build(story)
    return stream.getvalue()


def _location_label_cell(location: StorageLocation, target_width: float, target_height: float) -> Table:
    qr_size = min(34 * mm, target_height - 8 * mm)
    text_width = max(42 * mm, target_width - qr_size - 5 * mm)
    qr_image = _label_image(location.code, qr_size, qr_size)
    title_style = ParagraphStyle(
        "LocationTitle",
        fontName="Helvetica-Bold",
        fontSize=10,
        leading=12,
        textColor=colors.black,
    )
    code_style = ParagraphStyle(
        "LocationCode",
        fontName="Helvetica-Bold",
        fontSize=8,
        leading=10,
        textColor=colors.HexColor("#0066cc"),
        spaceAfter=3,
    )
    detail_style = ParagraphStyle(
        "LocationDetail",
        fontName="Helvetica",
        fontSize=7,
        leading=9,
        textColor=colors.HexColor("#333333"),
    )
    details = "<br/>".join(
        (
            f"Zone: <b>{escape(location.zone)}</b> &nbsp; Section: <b>{escape(location.section)}</b>",
            f"Rack: <b>{escape(location.rack)}</b> &nbsp; Shelf: <b>{escape(location.shelf)}</b>",
            f"Bin: <b>{escape(location.bin)}</b>",
        )
    )
    text_block = [
        Paragraph(escape(location.warehouse), title_style),
        Paragraph(escape(location.code), code_style),
        Paragraph(details, detail_style),
    ]
    label = Table(
        [[qr_image, text_block]],
        colWidths=[qr_size, text_width],
        rowHeights=[target_height],
    )
    label.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (0, 0), (0, 0), "CENTER"),
                ("LEFTPADDING", (0, 0), (0, 0), 0),
                ("RIGHTPADDING", (0, 0), (0, 0), 3 * mm),
                ("LEFTPADDING", (1, 0), (1, 0), 0),
                ("RIGHTPADDING", (1, 0), (1, 0), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ]
        )
    )
    return label


def location_labels_pdf(locations: list[StorageLocation]) -> bytes:
    stream = BytesIO()
    doc = SimpleDocTemplate(
        stream,
        pagesize=A4,
        rightMargin=8 * mm,
        leftMargin=8 * mm,
        topMargin=8 * mm,
        bottomMargin=8 * mm,
    )
    story = []
    rows_per_page = 5
    columns_per_page = 2
    labels_per_page = rows_per_page * columns_per_page
    col_width = 97 * mm
    row_height = 52 * mm

    for page_start in range(0, len(locations), labels_per_page):
        page_locations = locations[page_start:page_start + labels_per_page]
        table_rows = []
        for row_index in range(rows_per_page):
            row = []
            for column_index in range(columns_per_page):
                location_index = row_index * columns_per_page + column_index
                if location_index < len(page_locations):
                    row.append(
                        _location_label_cell(
                            page_locations[location_index],
                            col_width - 4 * mm,
                            row_height - 4 * mm,
                        )
                    )
                else:
                    row.append("")
            table_rows.append(row)
        table = Table(
            table_rows,
            colWidths=[col_width] * columns_per_page,
            rowHeights=[row_height] * rows_per_page,
        )
        table.setStyle(
            TableStyle(
                [
                    ("GRID", (0, 0), (-1, -1), 0.6, colors.lightgrey),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 2 * mm),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 2 * mm),
                    ("TOPPADDING", (0, 0), (-1, -1), 2 * mm),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 2 * mm),
                ]
            )
        )
        story.append(table)
        if page_start + labels_per_page < len(locations):
            story.append(PageBreak())
    if not story:
        styles = getSampleStyleSheet()
        story = [Paragraph("No storage locations selected", styles["BodyText"])]
    doc.build(story)
    return stream.getvalue()


def audit_report_pdf(batch: Batch) -> bytes:
    stream = BytesIO()
    doc = SimpleDocTemplate(stream, pagesize=A4, rightMargin=14 * mm, leftMargin=14 * mm, topMargin=14 * mm, bottomMargin=14 * mm)
    styles = getSampleStyleSheet()
    story = [
        Paragraph(f"Audit Report: {batch.batch_number}", styles["Title"]),
        Paragraph(f"Reference: {batch.party_name or '-'}", styles["BodyText"]),
        Paragraph(f"Status: {batch.status}", styles["BodyText"]),
        Spacer(1, 5 * mm),
    ]
    rows = [["Finding", "Serial", "Product", "Expected", "Scanned"]]
    for finding in batch.audit_findings:
        rows.append(
            [
                finding.finding_type,
                finding.serial_number,
                finding.product_name or "",
                finding.expected_status or "",
                finding.scanned_status or "",
            ]
        )
    table = Table(rows, repeatRows=1, colWidths=[28 * mm, 42 * mm, 58 * mm, 28 * mm, 28 * mm])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e8f2ff")),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.lightgrey),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
            ]
        )
    )
    story.append(table)
    doc.build(story)
    return stream.getvalue()
