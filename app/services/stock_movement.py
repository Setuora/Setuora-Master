from __future__ import annotations

from collections import Counter
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta, timezone
from io import BytesIO
from math import ceil, floor
from xml.sax.saxutils import escape

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy import desc, select
from sqlalchemy.orm import Session, selectinload

from app.models import AuditFinding, InventoryTransaction, Product, Serial, SerialStatus, TransactionType, WarehouseLevel
from app.services.audit import current_missing_stock_findings_query
from app.services.exports import safe_row, select_export_columns
from app.services.report_format import report_date
from app.services.settings import get_setting


STOCK_STATUSES = {SerialStatus.IN_STOCK.value, SerialStatus.RETURNED.value}
DEFAULT_ANALYSIS_DAYS = 90
DEFAULT_DEAD_BELOW = 10.0
DEFAULT_SLOW_BELOW = 40.0
DEFAULT_MEDIUM_UP_TO = 80.0
FRANCHISE_LEVELS = tuple(level.value for level in WarehouseLevel)


@dataclass(frozen=True)
class MovementConfig:
    analysis_days: int = DEFAULT_ANALYSIS_DAYS
    dead_below_pct: float = DEFAULT_DEAD_BELOW
    slow_below_pct: float = DEFAULT_SLOW_BELOW
    medium_up_to_pct: float = DEFAULT_MEDIUM_UP_TO


@dataclass(frozen=True)
class MovementFilters:
    product_id: int | None = None
    warehouse: str = ""
    category: str = ""
    brand: str = ""
    batch: str = ""
    franchise_level: str = ""
    expiry_period: str = ""
    movement: str = ""
    start: date | None = None
    end: date | None = None


def movement_config(db: Session) -> MovementConfig:
    try:
        config = MovementConfig(
            analysis_days=int(get_setting(db, "movement_analysis_days", str(DEFAULT_ANALYSIS_DAYS))),
            dead_below_pct=float(get_setting(db, "movement_dead_below_pct", str(DEFAULT_DEAD_BELOW))),
            slow_below_pct=float(get_setting(db, "movement_slow_below_pct", str(DEFAULT_SLOW_BELOW))),
            medium_up_to_pct=float(get_setting(db, "movement_medium_up_to_pct", str(DEFAULT_MEDIUM_UP_TO))),
        )
        validate_movement_config(config)
        return config
    except (TypeError, ValueError):
        return MovementConfig()


def validate_movement_config(config: MovementConfig) -> None:
    if config.analysis_days not in {30, 60, 90, 180, 365}:
        raise ValueError("Analysis period must be 30, 60, 90, 180, or 365 days.")
    cutoffs = (config.dead_below_pct, config.slow_below_pct, config.medium_up_to_pct)
    if any(value < 0 or value > 100 for value in cutoffs):
        raise ValueError("Movement thresholds must be between 0% and 100%.")
    if not config.dead_below_pct <= config.slow_below_pct <= config.medium_up_to_pct:
        raise ValueError("Thresholds must increase from dead to slow to medium.")


def analysis_window(
    config: MovementConfig,
    filters: MovementFilters,
    as_of: date | None = None,
) -> tuple[date, date, int]:
    end = filters.end or as_of or datetime.now(timezone.utc).date()
    start = filters.start or (end - timedelta(days=config.analysis_days - 1))
    if start > end:
        raise ValueError("Start date must be on or before end date.")
    return start, end, (end - start).days + 1


def warehouse_level(warehouse: str) -> str:
    text = warehouse.strip().upper().replace("-", " ")
    if "HOME FRANCHISE" in text:
        return "Home Franchise"
    if "TALUK" in text:
        return "Taluk Franchise"
    if "MASTER FRANCHISE" in text:
        return "Master Franchise"
    if "C&F" in text or "C AND F" in text or "CNF" in text:
        return "C&F"
    return "Company Warehouse"


def _serial_warehouse_level(serial: Serial | None) -> str:
    if not serial:
        return WarehouseLevel.COMPANY_WAREHOUSE.value
    try:
        return WarehouseLevel(serial.warehouse_level).value
    except (ValueError, TypeError):
        return warehouse_level(_warehouse(serial))


def movement_status(stock: int, sold: int, ratio_pct: float | None, config: MovementConfig) -> str:
    if stock == 0 and sold > 0:
        return "Fast Moving"
    if sold == 0 or ratio_pct is None:
        return "Dead Stock"
    if ratio_pct < config.dead_below_pct:
        return "Dead Stock"
    if ratio_pct < config.slow_below_pct:
        return "Slow Moving"
    if ratio_pct <= config.medium_up_to_pct:
        return "Medium Moving"
    return "Fast Moving"


def movement_css(status: str) -> str:
    return {
        "Fast Moving": "synced",
        "Medium Moving": "generated",
        "Slow Moving": "pending_sync",
        "Dead Stock": "failed",
    }.get(status, "")


def _warehouse(serial: Serial | None) -> str:
    return (serial.warehouse or "Main Warehouse").strip() if serial else "Main Warehouse"


def _product_matches(product: Product, filters: MovementFilters) -> bool:
    return (
        (not filters.product_id or product.id == filters.product_id)
        and (not filters.category or (product.category or "") == filters.category)
        and (not filters.brand or (product.brand or "") == filters.brand)
    )


def _serial_matches(serial: Serial, filters: MovementFilters) -> bool:
    warehouse = _warehouse(serial)
    return (
        (not filters.warehouse or warehouse == filters.warehouse)
        and (not filters.batch or (serial.product_batch_number or "") == filters.batch)
        and (not filters.franchise_level or _serial_warehouse_level(serial) == filters.franchise_level)
    )


def _expiry_matches(expiry_dates: list[date], expiry_period: str, as_of: date) -> bool:
    if not expiry_period:
        return True
    if expiry_period == "none":
        return not expiry_dates
    if expiry_period == "expired":
        return any(expiry < as_of for expiry in expiry_dates)
    try:
        days = int(expiry_period)
    except ValueError:
        return True
    deadline = as_of + timedelta(days=days)
    return any(as_of <= expiry <= deadline for expiry in expiry_dates)


def _expiry_date_matches(expiry: date | None, expiry_period: str, as_of: date) -> bool:
    if not expiry_period:
        return True
    if expiry_period == "none":
        return expiry is None
    if expiry is None:
        return False
    if expiry_period == "expired":
        return expiry < as_of
    try:
        days = int(expiry_period)
    except ValueError:
        return True
    return as_of <= expiry <= as_of + timedelta(days=days)


def _inventory_signal(stock: int, sold: int, daily_sales: float, clear_days: float | None, analysis_days: int) -> str:
    if stock == 0 and sold > 0:
        return "Understocked"
    if stock > 0 and daily_sales == 0:
        return "Overstocked"
    if clear_days is not None and clear_days < 14:
        return "Understocked"
    if clear_days is not None and clear_days > analysis_days * 2:
        return "Overstocked"
    return "Normal Movement"


def _suggested_action(
    status: str,
    signal: str,
    expiry_risk: str,
    estimated_unsold: int,
) -> str:
    if expiry_risk == "High Expiry Risk":
        return f"Discount or bundle; {estimated_unsold} units at risk"
    if expiry_risk == "Watch":
        return "Increase promotion and monitor weekly"
    if signal == "Understocked":
        return "Review reorder quantity"
    if status == "Dead Stock":
        return "Stop purchasing; return or liquidate"
    if signal == "Overstocked":
        return "Transfer stock and reduce next purchase"
    if status == "Slow Moving":
        return "Promote, bundle, or offer a discount"
    if status == "Fast Moving":
        return "Maintain availability"
    return "Monitor movement"


def _expiry_projection(
    expiry_quantities: Counter[date],
    daily_sales: float,
    as_of: date,
) -> tuple[str, str, int, date | None, float | None]:
    if not expiry_quantities:
        return "Not applicable", "", 0, None, None
    nearest = min(expiry_quantities)
    nearest_days = (nearest - as_of).days
    cumulative = 0
    estimated_unsold = 0
    for expiry in sorted(expiry_quantities):
        cumulative += expiry_quantities[expiry]
        available_days = max((expiry - as_of).days, 0)
        projected_sales = floor(daily_sales * available_days)
        estimated_unsold = max(estimated_unsold, cumulative - projected_sales)
    if estimated_unsold > 0:
        return "High Expiry Risk", "failed", estimated_unsold, nearest, nearest_days / 30
    expiring_stock = sum(expiry_quantities.values())
    days_to_sell_expiring = expiring_stock / daily_sales if daily_sales else None
    if days_to_sell_expiring is not None and nearest_days > 0 and days_to_sell_expiring >= nearest_days * 0.8:
        return "Watch", "pending_sync", 0, nearest, nearest_days / 30
    return "Safe", "synced", 0, nearest, nearest_days / 30


def stock_movement_rows(
    db: Session,
    config: MovementConfig | None = None,
    filters: MovementFilters | None = None,
    as_of: date | None = None,
) -> tuple[list[dict[str, object]], dict[str, object]]:
    config = config or movement_config(db)
    filters = filters or MovementFilters()
    as_of = as_of or filters.end or datetime.now(timezone.utc).date()
    start, end, analysis_days = analysis_window(config, filters, as_of=as_of)
    start_at = datetime.combine(start, time.min, tzinfo=timezone.utc)
    end_at = datetime.combine(end + timedelta(days=1), time.min, tzinfo=timezone.utc)

    stock_groups: dict[tuple[int, str, str], dict[str, object]] = {}
    stock_serials = db.scalars(
        select(Serial)
        .where(Serial.active.is_(True), Serial.status.in_(STOCK_STATUSES))
        .options(selectinload(Serial.product))
    ).all()
    for serial in stock_serials:
        if not _product_matches(serial.product, filters) or not _serial_matches(serial, filters):
            continue
        if not _expiry_date_matches(serial.expiry_date, filters.expiry_period, as_of):
            continue
        warehouse = _warehouse(serial)
        franchise_level = _serial_warehouse_level(serial)
        key = (serial.product_id, warehouse, franchise_level)
        group = stock_groups.setdefault(
            key,
            {
                "product": serial.product,
                "warehouse": warehouse,
                "franchise_level": franchise_level,
                "stock": 0,
                "batches": set(),
                "expiry_quantities": Counter(),
            },
        )
        group["stock"] = int(group["stock"]) + 1
        if serial.product_batch_number:
            group["batches"].add(serial.product_batch_number)  # type: ignore[union-attr]
        if serial.expiry_date:
            group["expiry_quantities"][serial.expiry_date] += 1  # type: ignore[index]

    sales_groups: Counter[tuple[int, str, str]] = Counter()
    sold_transactions = db.scalars(
        select(InventoryTransaction)
        .where(
            InventoryTransaction.transaction_type == TransactionType.SALE.value,
            InventoryTransaction.status_to == SerialStatus.SOLD.value,
            InventoryTransaction.created_at >= start_at,
            InventoryTransaction.created_at < end_at,
        )
        .options(
            selectinload(InventoryTransaction.product),
            selectinload(InventoryTransaction.serial),
        )
    ).all()
    products_by_key: dict[tuple[int, str, str], Product] = {}
    for transaction in sold_transactions:
        if not transaction.product or not transaction.serial:
            continue
        if not _product_matches(transaction.product, filters) or not _serial_matches(transaction.serial, filters):
            continue
        warehouse = _warehouse(transaction.serial)
        key = (transaction.product.id, warehouse, _serial_warehouse_level(transaction.serial))
        sales_groups[key] += 1
        products_by_key[key] = transaction.product

    rows: list[dict[str, object]] = []
    for key in stock_groups.keys() | sales_groups.keys():
        group = stock_groups.get(key)
        product = group["product"] if group else products_by_key[key]
        warehouse = str(group["warehouse"] if group else key[1])
        franchise_level = str(group["franchise_level"] if group else key[2])
        stock = int(group["stock"]) if group else 0
        sold = int(sales_groups[key])
        expiry_quantities: Counter[date] = group["expiry_quantities"] if group else Counter()  # type: ignore[assignment]
        expiry_dates = list(expiry_quantities)
        if not _expiry_matches(expiry_dates, filters.expiry_period, as_of):
            continue
        ratio_pct = (sold / stock * 100) if stock else None
        daily_sales = sold / analysis_days
        clear_days = stock / daily_sales if daily_sales else None
        status = movement_status(stock, sold, ratio_pct, config)
        if filters.movement and status != filters.movement:
            continue
        signal = _inventory_signal(stock, sold, daily_sales, clear_days, analysis_days)
        expiry_risk, expiry_css, estimated_unsold, nearest_expiry, months_remaining = _expiry_projection(
            expiry_quantities,
            daily_sales,
            as_of,
        )
        batches: set[str] = group["batches"] if group else set()  # type: ignore[assignment]
        batch_label = ", ".join(sorted(batches)) if len(batches) <= 2 else f"Multiple ({len(batches)})"
        rows.append(
            {
                "product_id": product.id,
                "product_code": product.product_code,
                "product_name": product.product_name,
                "category": product.category or "",
                "brand": product.brand or "",
                "warehouse": warehouse,
                "franchise_level": franchise_level,
                "batch": batch_label or "-",
                "current_stock": stock,
                "units_sold": sold,
                "movement_ratio": round(ratio_pct, 1) if ratio_pct is not None else None,
                "average_daily_sales": round(daily_sales, 2),
                "average_weekly_sales": round(daily_sales * 7, 2),
                "average_monthly_sales": round(daily_sales * 30, 2),
                "estimated_days": round(clear_days, 1) if clear_days is not None else None,
                "estimated_months": round(clear_days / 30, 1) if clear_days is not None else None,
                "movement_status": status,
                "movement_css": movement_css(status),
                "inventory_signal": signal,
                "nearest_expiry": nearest_expiry,
                "months_remaining": round(months_remaining, 1) if months_remaining is not None else None,
                "expiry_risk": expiry_risk,
                "expiry_css": expiry_css,
                "estimated_unsold": estimated_unsold,
                "suggested_action": _suggested_action(status, signal, expiry_risk, estimated_unsold),
            }
        )

    status_order = {"Dead Stock": 0, "Slow Moving": 1, "Medium Moving": 2, "Fast Moving": 3}
    rows.sort(
        key=lambda row: (
            row["expiry_risk"] != "High Expiry Risk",
            status_order.get(str(row["movement_status"]), 9),
            str(row["product_name"]),
            str(row["warehouse"]),
        )
    )
    movement_counts = Counter(str(row["movement_status"]) for row in rows)
    summary: dict[str, object] = {
        "start": start,
        "end": end,
        "analysis_days": analysis_days,
        "total_products": len({row["product_id"] for row in rows}),
        "current_stock": sum(int(row["current_stock"]) for row in rows),
        "units_sold": sum(int(row["units_sold"]) for row in rows),
        "fast": movement_counts["Fast Moving"],
        "medium": movement_counts["Medium Moving"],
        "slow": movement_counts["Slow Moving"],
        "dead": movement_counts["Dead Stock"],
        "expiry_risk": sum(row["expiry_risk"] == "High Expiry Risk" for row in rows),
        "overstocked": sum(row["inventory_signal"] == "Overstocked" for row in rows),
        "understocked": sum(row["inventory_signal"] == "Understocked" for row in rows),
    }
    return rows, summary


def product_inventory_metrics(
    db: Session,
    products: list[Product],
    config: MovementConfig | None = None,
    as_of: date | None = None,
) -> tuple[dict[int, dict[str, object]], int]:
    """Summarize sales, physically available stock, and restock timing per product."""
    as_of = as_of or datetime.now(timezone.utc).date()
    config = config or movement_config(db)
    movement_rows, summary = stock_movement_rows(db, config=config, as_of=as_of)
    analysis_days = int(summary["analysis_days"])
    metrics: dict[int, dict[str, object]] = {
        product.id: {
            "units_sold": 0,
            "system_stock": 0,
            "available_stock": 0,
            "missing_stock": 0,
            "restock_label": "Not forecast",
            "restock_detail": f"No sales in {analysis_days} days",
            "restock_css": "generated",
        }
        for product in products
    }

    for row in movement_rows:
        product_id = int(row["product_id"])
        if product_id not in metrics:
            continue
        metrics[product_id]["units_sold"] = int(metrics[product_id]["units_sold"]) + int(row["units_sold"])
        metrics[product_id]["system_stock"] = int(metrics[product_id]["system_stock"]) + int(row["current_stock"])

    current_missing_findings = db.scalars(
        current_missing_stock_findings_query().options(selectinload(AuditFinding.serial))
    ).all()
    for finding in current_missing_findings:
        serial = finding.serial
        if serial and serial.product_id in metrics:
            metrics[serial.product_id]["missing_stock"] = int(metrics[serial.product_id]["missing_stock"]) + 1

    for metric in metrics.values():
        units_sold = int(metric["units_sold"])
        system_stock = int(metric["system_stock"])
        missing_stock = int(metric["missing_stock"])
        available_stock = max(system_stock - missing_stock, 0)
        metric["available_stock"] = available_stock
        if not units_sold:
            continue
        if not available_stock:
            metric.update(
                {
                    "restock_label": "Restock now",
                    "restock_detail": "No available stock",
                    "restock_css": "failed",
                }
            )
            continue
        daily_sales = units_sold / analysis_days
        days_remaining = max(1, ceil(available_stock / daily_sales))
        metric.update(
            {
                "restock_label": f"In {days_remaining} days",
                "restock_detail": f"By {report_date(as_of + timedelta(days=days_remaining))}",
                "restock_css": "pending_sync" if days_remaining <= 30 else "active",
            }
        )
    return metrics, analysis_days


def product_sales_transactions(
    db: Session,
    product_ids: set[int],
    analysis_days: int,
    as_of: date | None = None,
) -> list[InventoryTransaction]:
    if not product_ids:
        return []
    as_of = as_of or datetime.now(timezone.utc).date()
    start = as_of - timedelta(days=analysis_days - 1)
    start_at = datetime.combine(start, time.min, tzinfo=timezone.utc)
    end_at = datetime.combine(as_of + timedelta(days=1), time.min, tzinfo=timezone.utc)
    return db.scalars(
        select(InventoryTransaction)
        .where(
            InventoryTransaction.product_id.in_(product_ids),
            InventoryTransaction.transaction_type == TransactionType.SALE.value,
            InventoryTransaction.status_to == SerialStatus.SOLD.value,
            InventoryTransaction.created_at >= start_at,
            InventoryTransaction.created_at < end_at,
        )
        .order_by(desc(InventoryTransaction.created_at))
        .options(
            selectinload(InventoryTransaction.user),
            selectinload(InventoryTransaction.batch),
        )
    ).all()


def product_sales_report_pdf(
    product: Product,
    metric: dict[str, object],
    sales: list[InventoryTransaction],
    analysis_days: int,
    as_of: date | None = None,
) -> bytes:
    as_of = as_of or datetime.now(timezone.utc).date()
    stream = BytesIO()
    doc = SimpleDocTemplate(
        stream,
        pagesize=A4,
        leftMargin=14 * mm,
        rightMargin=14 * mm,
        topMargin=14 * mm,
        bottomMargin=14 * mm,
    )
    styles = getSampleStyleSheet()
    story = [
        Paragraph(escape(f"Product Sales Report: {product.product_code} - {product.product_name}"), styles["Title"]),
        Paragraph(
            escape(
                f"Period: {report_date(as_of - timedelta(days=analysis_days - 1))} to {report_date(as_of)} "
                f"({analysis_days} days)"
            ),
            styles["BodyText"],
        ),
        Spacer(1, 5 * mm),
    ]
    summary = Table(
        [
            ["Sales", "Available stock", "Missing stock", "Restock"],
            [
                metric["units_sold"],
                metric["available_stock"],
                metric["missing_stock"],
                f"{metric['restock_label']} - {metric['restock_detail']}",
            ],
        ],
        colWidths=[32 * mm, 38 * mm, 34 * mm, 76 * mm],
    )
    summary.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e8f2ff")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.lightgrey),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    story.extend([summary, Spacer(1, 7 * mm), Paragraph("Individual sales", styles["Heading2"])])
    sale_rows: list[list[object]] = [["Date", "Serial", "Sold by", "Batch / reference", "Tally reference"]]
    for sale in sales:
        reference = sale.reference_number or (sale.batch.batch_number if sale.batch else "")
        sale_rows.append(
            [
                report_date(sale.created_at),
                sale.serial_number or "",
                sale.user.username,
                reference,
                sale.tally_reference or "",
            ]
        )
    if not sales:
        sale_rows.append(["No sales during this period", "", "", "", ""])
    sales_table = Table(
        sale_rows,
        repeatRows=1,
        colWidths=[32 * mm, 42 * mm, 28 * mm, 40 * mm, 38 * mm],
    )
    sales_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#202124")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.lightgrey),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f7f7f7")]),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    story.append(sales_table)
    doc.build(story)
    return stream.getvalue()


MOVEMENT_EXPORT_HEADERS = [
    "Product Code",
    "Product Name",
    "Category",
    "Brand",
    "Warehouse",
    "Franchise Level",
    "Batch",
    "Current Stock",
    "Units Sold",
    "Movement Ratio %",
    "Average Daily Sales",
    "Average Weekly Sales",
    "Average Monthly Sales",
    "Estimated Days to Sell",
    "Estimated Months to Sell",
    "Movement Status",
    "Inventory Signal",
    "Nearest Expiry",
    "Months Remaining",
    "Expiry Risk",
    "Estimated Unsold",
    "Suggested Action",
]


def movement_export_row(row: dict[str, object]) -> list[object]:
    return safe_row(
        [
            row["product_code"],
            row["product_name"],
            row["category"],
            row["brand"],
            row["warehouse"],
            row["franchise_level"],
            row["batch"],
            row["current_stock"],
            row["units_sold"],
            row["movement_ratio"] if row["movement_ratio"] is not None else "",
            row["average_daily_sales"],
            row["average_weekly_sales"],
            row["average_monthly_sales"],
            row["estimated_days"] if row["estimated_days"] is not None else "No sales",
            row["estimated_months"] if row["estimated_months"] is not None else "No sales",
            row["movement_status"],
            row["inventory_signal"],
            report_date(row["nearest_expiry"]) if row["nearest_expiry"] else "",
            row["months_remaining"] if row["months_remaining"] is not None else "",
            row["expiry_risk"],
            row["estimated_unsold"],
            row["suggested_action"],
        ]
    )


def stock_movement_xlsx(
    rows: list[dict[str, object]],
    summary: dict[str, object],
    fields: list[str] | None = None,
) -> bytes:
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Stock Movement"
    sheet.append(
        [
            "Analysis Period",
            f"{report_date(summary['start'])} to {report_date(summary['end'])} ({summary['analysis_days']} days)",
        ]
    )
    sheet.append([])
    export_rows = [movement_export_row(row) for row in rows]
    headers, export_rows = select_export_columns(MOVEMENT_EXPORT_HEADERS, export_rows, fields)
    sheet.append(headers)
    for row in export_rows:
        sheet.append(row)
    sheet.freeze_panes = "A4"
    sheet.auto_filter.ref = f"A3:{sheet.cell(3, sheet.max_column).coordinate}"
    for column in sheet.columns:
        width = max(len(str(cell.value or "")) for cell in column)
        sheet.column_dimensions[column[0].column_letter].width = min(max(width + 2, 12), 42)
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def stock_movement_pdf(rows: list[dict[str, object]], summary: dict[str, object]) -> bytes:
    stream = BytesIO()
    doc = SimpleDocTemplate(
        stream,
        pagesize=landscape(A4),
        leftMargin=9 * mm,
        rightMargin=9 * mm,
        topMargin=9 * mm,
        bottomMargin=9 * mm,
    )
    styles = getSampleStyleSheet()
    story = [
        Paragraph("Stock Movement & Slow Moving Inventory", styles["Title"]),
        Paragraph(
            escape(
                "Analysis period: "
                f"{report_date(summary['start'])} to {report_date(summary['end'])} "
                f"({summary['analysis_days']} days)"
            ),
            styles["BodyText"],
        ),
        Spacer(1, 4 * mm),
    ]
    headers = ["Product", "Warehouse", "Stock", "Sold", "Ratio", "Monthly", "Sell-out", "Movement", "Expiry risk", "Action"]
    data: list[list[object]] = [headers]
    small = styles["BodyText"]
    small.fontSize = 7
    small.leading = 8
    for row in rows:
        data.append(
            [
                Paragraph(escape(f"{row['product_code']} - {row['product_name']}"), small),
                Paragraph(escape(str(row["warehouse"])), small),
                row["current_stock"],
                row["units_sold"],
                f"{row['movement_ratio']}%" if row["movement_ratio"] is not None else "-",
                row["average_monthly_sales"],
                f"{row['estimated_months']} mo" if row["estimated_months"] is not None else "No sales",
                Paragraph(escape(str(row["movement_status"])), small),
                Paragraph(
                    escape(
                        f"{row['expiry_risk']}"
                        + (f" ({row['estimated_unsold']} unsold)" if row["estimated_unsold"] else "")
                    ),
                    small,
                ),
                Paragraph(escape(str(row["suggested_action"])), small),
            ]
        )
    table = Table(
        data,
        repeatRows=1,
        colWidths=[43 * mm, 28 * mm, 13 * mm, 13 * mm, 15 * mm, 17 * mm, 18 * mm, 24 * mm, 32 * mm, 58 * mm],
    )
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#202124")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#d9d9d9")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f7f7f7")]),
                ("LEFTPADDING", (0, 0), (-1, -1), 3),
                ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )
    story.append(table)
    doc.build(story)
    return stream.getvalue()
