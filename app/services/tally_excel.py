from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.orm import Session

from app.models import Batch, BatchItem, BatchStatus, BatchType, GstRegistrationType, ScanLog, Serial, User
from app.services.assignment import AssignmentLine, MAX_ASSIGNMENT_QUANTITY, parse_bulk_assignment_xlsx
from app.services.expiry import fefo_available_statuses, fefo_candidate_serials
from app.services.exports import safe_row, select_export_columns
from app.services.report_format import EXCEL_DATE_FORMAT, batch_voucher_number
from app.services.inventory import InventoryError
from app.services.settings import gst_rate_key, parse_sales_gst_ledger_mappings
from app.services.tally import (
    DEFAULT_PURCHASE_VOUCHER_TYPE,
    DEFAULT_SALES_RETURN_VOUCHER_TYPE,
    DEFAULT_SALES_VOUCHER_TYPE,
    TallySyncError,
)
from app.services.voucher import calculate_voucher_summary


MAX_TALLY_EXCEL_UPLOAD_BYTES = 5 * 1024 * 1024
TALLY_EXCEL_IMPORT_BATCH_TYPES = {
    BatchType.SALE.value,
    BatchType.ISSUE.value,
    BatchType.PURCHASE_RETURN.value,
}
TALLY_EXCEL_EXPORT_BATCH_TYPES = {
    BatchType.PURCHASE.value,
    BatchType.RECEIVE.value,
    BatchType.SALE.value,
    BatchType.SALES_RETURN.value,
    BatchType.PURCHASE_RETURN.value,
    BatchType.ISSUE.value,
}
TALLY_ACCOUNTING_VOUCHER_BATCH_TYPES = {
    BatchType.PURCHASE.value,
    BatchType.RECEIVE.value,
    BatchType.SALE.value,
    BatchType.SALES_RETURN.value,
}
TALLY_ACCOUNTING_VOUCHER_SHEET = "Accounting Voucher"
TALLY_ACCOUNTING_VOUCHER_HEADERS = [
    "Voucher Date",
    "Voucher Type Name",
    "Voucher Number",
    "Buyer/Supplier - Address",
    "Buyer/Supplier - Pincode",
    "Ledger Name",
    "Ledger Amount",
    "Ledger Amount Dr/Cr",
    "Item Name",
    "Billed Quantity",
    "Item Rate",
    "Item Rate per",
    "Disc%",
    "Item Amount",
    "Change Mode ",
    "Buyer/Supplier - GST Registration Type",
    "Buyer/Supplier - GSTIN/UIN",
    "Buyer/Supplier - Place of Supply",
    "HSN/SAC Details",
    "HSN/SAC",
    "GST Rate Details",
    "GST Taxability Type",
    "GST Nature of Transaction",
    "IGST Rate",
    "CGST Rate",
    "SGST/UTGST Rate",
    "Taxable Value",
]
TALLY_ACCOUNTING_REQUIRED_EXPORT_FIELDS = [
    "Voucher Date",
    "Voucher Type Name",
    "Voucher Number",
    "Ledger Name",
    "Ledger Amount",
    "Item Name",
    "Billed Quantity",
    "Item Rate",
    "Item Amount",
    "Change Mode",
    "Buyer/Supplier - Place of Supply",
    "HSN/SAC",
    "GST Nature of Transaction",
    "Taxable Value",
]
TALLY_ITEM_SUMMARY_HEADERS = [
    "Sl",
    "Description of Goods",
    "Product Code",
    "Tally Stock Item",
    "HSN/SAC",
    "Quantity",
    "Unit",
    "Rate",
    "Discount %",
    "GST %",
    "CGST %",
    "SGST %",
    "IGST %",
    "Taxable Value",
    "CGST Amount",
    "SGST Amount",
    "IGST Amount",
    "Amount excl. GST",
]
_IST = timezone(timedelta(hours=5, minutes=30))


@dataclass(frozen=True)
class TallyExcelImportResult:
    product_lines: int
    quantity: int


def tally_accounting_default_deselected_fields(batch: Batch) -> list[str]:
    if batch.batch_type not in TALLY_ACCOUNTING_VOUCHER_BATCH_TYPES:
        return []
    batch_type = BatchType(batch.batch_type)
    if batch_type not in {BatchType.SALE, BatchType.SALES_RETURN}:
        return []

    registration_type = (batch.party_gst_registration_type or "").strip()
    if not registration_type:
        registration_type = GstRegistrationType.UNREGISTERED_CONSUMER.value
    gst_treatment = (batch.gst_treatment or "").strip().upper()
    deselected: set[str] = set()

    if registration_type == GstRegistrationType.UNREGISTERED_CONSUMER.value:
        deselected.update(
            {
                "Buyer/Supplier - Address",
                "Buyer/Supplier - Pincode",
                "Buyer/Supplier - GST Registration Type",
                "Buyer/Supplier - GSTIN/UIN",
                "HSN/SAC Details",
            }
        )

    if gst_treatment == "INTER_STATE":
        deselected.update({"CGST Rate", "SGST/UTGST Rate"})
    else:
        deselected.add("IGST Rate")

    if registration_type == GstRegistrationType.COMPOSITION.value:
        deselected.update(
            {
                "GST Rate Details",
                "GST Taxability Type",
                "IGST Rate",
                "CGST Rate",
                "SGST/UTGST Rate",
            }
        )

    deselected.difference_update(TALLY_ACCOUNTING_REQUIRED_EXPORT_FIELDS)
    return _ordered_accounting_fields(deselected)


def batch_tally_xlsx(
    batch: Batch,
    settings: dict[str, str] | None = None,
    fields: list[str] | None = None,
    overrides: dict[str, str] | None = None,
) -> bytes:
    if batch.batch_type in TALLY_ACCOUNTING_VOUCHER_BATCH_TYPES:
        return _accounting_voucher_xlsx(batch, settings or {}, fields, overrides or {})
    return _item_summary_xlsx(batch, fields, overrides or {})


def _accounting_voucher_xlsx(
    batch: Batch,
    settings: dict[str, str],
    fields: list[str] | None = None,
    overrides: dict[str, str] | None = None,
) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = TALLY_ACCOUNTING_VOUCHER_SHEET

    export_rows = _accounting_voucher_rows(batch, settings, overrides or {})
    headers, rows = select_export_columns(export_rows[0], export_rows[1:], fields)
    for row in [headers, *rows]:
        sheet.append(safe_row(row))

    _format_accounting_sheet(sheet)

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _accounting_voucher_rows(
    batch: Batch,
    settings: dict[str, str],
    overrides: dict[str, str] | None = None,
) -> list[list[object]]:
    overrides = overrides or {}
    summary = calculate_voucher_summary(batch)
    batch_type = BatchType(batch.batch_type)
    is_sale = batch_type == BatchType.SALE
    is_sales_side = batch_type in {BatchType.SALE, BatchType.SALES_RETURN}
    voucher_type = overrides.get("voucher_type", "").strip() or _voucher_type(settings, batch_type)
    change_mode = "Accounting Invoice" if is_sales_side else "As Voucher"
    common = {
        "Voucher Date": _voucher_date(batch),
        "Voucher Type Name": voucher_type,
        "Voucher Number": batch_voucher_number(batch, overrides.get("voucher_number", "")),
        "Buyer/Supplier - Address": "",
        "Buyer/Supplier - Pincode": "",
        "Change Mode ": change_mode,
        **_party_gst_fields(batch, is_sales_side),
    }
    rows: list[dict[str, object]] = []

    party_name = overrides.get("party_ledger", "").strip() or (batch.party_name or "").strip()
    party_signed_amount = -summary.final_value if is_sale else summary.final_value
    rows.append(_posting_row(common, party_name, party_signed_amount))

    sales_mappings = _sales_gst_mappings(settings)
    purchase_ledgers = _purchase_ledgers(settings)
    tax_postings: dict[str, Decimal] = {}

    for line in summary.lines:
        if is_sales_side:
            ledgers = _sales_ledgers(
                settings,
                sales_mappings,
                line.gst_rate,
                line.cgst_rate,
                line.sgst_rate,
                line.igst_rate,
            )
            item_ledger = ledgers["sales"]
            tax_sign = Decimal("1") if is_sale else Decimal("-1")
            if line.cgst_amount:
                tax_postings[ledgers["cgst"]] = (
                    tax_postings.get(ledgers["cgst"], Decimal("0")) + line.cgst_amount * tax_sign
                )
            if line.sgst_amount:
                tax_postings[ledgers["sgst"]] = (
                    tax_postings.get(ledgers["sgst"], Decimal("0")) + line.sgst_amount * tax_sign
                )
            if line.igst_amount:
                tax_postings[ledgers["igst"]] = (
                    tax_postings.get(ledgers["igst"], Decimal("0")) + line.igst_amount * tax_sign
                )
            item_signed_amount = line.taxable_value if is_sale else -line.taxable_value
        else:
            item_ledger = purchase_ledgers["purchase"]
            item_signed_amount = -line.taxable_value

        rows.append(
            _posting_row(
                common,
                item_ledger,
                item_signed_amount,
                item={
                    "Item Name": line.tally_stock_item_name or line.product_name,
                    "Billed Quantity": line.quantity,
                    "Item Rate": line.rate,
                    "Item Rate per": line.unit,
                    "Disc%": line.discount_rate if line.discount_rate else "",
                    "Item Amount": line.taxable_value,
                    "HSN/SAC Details": "Specify Details Here" if line.hsn else "",
                    "HSN/SAC": line.hsn,
                    "GST Rate Details": "Specify Details Here",
                    "GST Taxability Type": "Taxable",
                    "GST Nature of Transaction": _gst_nature(batch, is_sales_side),
                    "IGST Rate": line.igst_rate,
                    "CGST Rate": line.cgst_rate,
                    "SGST/UTGST Rate": line.sgst_rate,
                    "Taxable Value": line.taxable_value,
                },
            )
        )

    if not is_sales_side:
        if summary.cgst_amount:
            tax_postings[purchase_ledgers["cgst"]] = (
                tax_postings.get(purchase_ledgers["cgst"], Decimal("0")) - summary.cgst_amount
            )
        if summary.sgst_amount:
            tax_postings[purchase_ledgers["sgst"]] = (
                tax_postings.get(purchase_ledgers["sgst"], Decimal("0")) - summary.sgst_amount
            )

    for ledger_name, signed_amount in tax_postings.items():
        if signed_amount:
            rows.append(_posting_row(common, ledger_name, signed_amount))

    if summary.round_off:
        round_off_ledger = settings.get("round_off_ledger_name", "").strip() or "ROUND OFF"
        round_off_signed_amount = summary.round_off if is_sale else -summary.round_off
        rows.append(_posting_row(common, round_off_ledger, round_off_signed_amount))

    headers = _accounting_voucher_headers()
    return [headers] + [[row.get(header, "") for header in headers] for row in rows]


def _accounting_voucher_headers() -> list[str]:
    return list(TALLY_ACCOUNTING_VOUCHER_HEADERS)


def _ordered_accounting_fields(fields: set[str]) -> list[str]:
    return [header.strip() for header in TALLY_ACCOUNTING_VOUCHER_HEADERS if header.strip() in fields]


def _item_summary_xlsx(
    batch: Batch,
    fields: list[str] | None = None,
    overrides: dict[str, str] | None = None,
) -> bytes:
    overrides = overrides or {}
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Tally Voucher"
    summary = calculate_voucher_summary(batch)

    sheet.append(["Voucher Type", overrides.get("voucher_type", "").strip() or batch.batch_type])
    sheet.append(["Voucher Number", batch_voucher_number(batch, overrides.get("voucher_number", ""))])
    sheet.append(["Party Ledger", overrides.get("party_ledger", "").strip() or batch.party_name or ""])
    sheet.append(["Date", _voucher_date(batch)])
    sheet["B4"].number_format = EXCEL_DATE_FORMAT
    sheet.append([])
    total_quantity = 0
    rows: list[list[object]] = []
    for index, line in enumerate(summary.lines, start=1):
        total_quantity += line.quantity
        rows.append(
            safe_row(
                [
                    index,
                    line.tally_stock_item_name or line.product_name,
                    line.product_code,
                    line.tally_stock_item_name,
                    line.hsn,
                    line.quantity,
                    line.unit,
                    float(line.rate),
                    float(line.discount_rate),
                    float(line.gst_rate),
                    float(line.cgst_rate),
                    float(line.sgst_rate),
                    float(line.igst_rate),
                    float(line.taxable_value),
                    float(line.cgst_amount),
                    float(line.sgst_amount),
                    float(line.igst_amount),
                    float(line.taxable_value),
                ]
            )
        )

    rows.append(
        safe_row(
            [
                "",
                "Total",
                "",
                "",
                "",
                total_quantity,
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                float(summary.taxable_value),
                float(summary.cgst_amount),
                float(summary.sgst_amount),
                float(summary.igst_amount),
                float(summary.taxable_value),
            ]
        )
    )
    headers, rows = select_export_columns(TALLY_ITEM_SUMMARY_HEADERS, rows, fields)
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    sheet.freeze_panes = "A7"
    sheet.auto_filter.ref = f"A6:{sheet.cell(6, sheet.max_column).coordinate}"
    _autosize(sheet)

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _voucher_date(batch: Batch):
    moment = batch.submitted_at or batch.created_at or datetime.now(timezone.utc)
    if moment.tzinfo is None:
        moment = moment.replace(tzinfo=timezone.utc)
    return moment.astimezone(_IST).date()


def _voucher_type(settings: dict[str, str], batch_type: BatchType) -> str:
    if batch_type == BatchType.SALE:
        return settings.get("sales_voucher_type", "").strip() or DEFAULT_SALES_VOUCHER_TYPE
    if batch_type == BatchType.SALES_RETURN:
        return settings.get("sales_return_voucher_type", "").strip() or DEFAULT_SALES_RETURN_VOUCHER_TYPE
    return settings.get("purchase_voucher_type", "").strip() or DEFAULT_PURCHASE_VOUCHER_TYPE


def _party_gst_fields(batch: Batch, is_sales_side: bool) -> dict[str, object]:
    registration_type = (batch.party_gst_registration_type or "").strip()
    if is_sales_side and not registration_type:
        registration_type = GstRegistrationType.UNREGISTERED_CONSUMER.value
    return {
        "Buyer/Supplier - GST Registration Type": registration_type,
        "Buyer/Supplier - GSTIN/UIN": (batch.party_gstin or "").strip(),
        "Buyer/Supplier - Place of Supply": (batch.party_state or "").strip(),
    }


def _gst_nature(batch: Batch, is_sales_side: bool) -> str:
    if not is_sales_side:
        return ""
    if (getattr(batch, "gst_treatment", "") or "").upper() == "INTER_STATE":
        return "Interstate Sales - Taxable"
    return "Local Sales - Taxable"


def _sales_gst_mappings(settings: dict[str, str]) -> dict[str, dict[str, str]]:
    try:
        return parse_sales_gst_ledger_mappings(settings.get("sales_gst_ledger_mappings"))
    except ValueError as exc:
        raise TallySyncError(str(exc), retryable=False) from exc


def _sales_ledgers(
    settings: dict[str, str],
    mappings: dict[str, dict[str, str]],
    gst_rate: Decimal,
    cgst_rate: Decimal,
    sgst_rate: Decimal,
    igst_rate: Decimal,
) -> dict[str, str]:
    key = gst_rate_key(gst_rate)
    if key in mappings:
        return mappings[key]
    return {
        "sales": settings.get("sales_ledger_name", "").strip() or f"Sales @ {key}%",
        "cgst": settings.get("cgst_ledger_name", "").strip() or f"Output CGST @ {gst_rate_key(cgst_rate)}%",
        "sgst": settings.get("sgst_ledger_name", "").strip() or f"Output SGST @ {gst_rate_key(sgst_rate)}%",
        "igst": settings.get("igst_ledger_name", "").strip() or f"Output IGST @ {gst_rate_key(igst_rate or gst_rate)}%",
    }


def _purchase_ledgers(settings: dict[str, str]) -> dict[str, str]:
    return {
        "purchase": settings.get("purchase_ledger_name", "").strip() or "Purchase",
        "cgst": settings.get("cgst_ledger_name", "").strip() or "Input CGST",
        "sgst": settings.get("sgst_ledger_name", "").strip() or "Input SGST",
    }


def _posting_row(
    common: dict[str, object],
    ledger_name: str,
    signed_amount: Decimal,
    *,
    item: dict[str, object] | None = None,
) -> dict[str, object]:
    row = dict(common)
    amount = Decimal(signed_amount)
    row.update(
        {
            "Ledger Name": ledger_name,
            "Ledger Amount": _number(abs(amount)),
            "Ledger Amount Dr/Cr": "Cr" if amount >= 0 else "Dr",
        }
    )
    if item:
        row.update({key: _number(value) if isinstance(value, Decimal) else value for key, value in item.items()})
    return row


def _number(value: object) -> object:
    if value in {None, ""}:
        return ""
    amount = Decimal(str(value)).quantize(Decimal("0.01"))
    if amount == amount.to_integral_value():
        return int(amount)
    return float(amount)


def _format_accounting_sheet(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="FEC530")
    header_font = Font(bold=True, color="000000")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{sheet.cell(1, sheet.max_column).coordinate}"
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            header = sheet.cell(1, cell.column).value
            if header == "Voucher Date":
                cell.number_format = EXCEL_DATE_FORMAT
            elif header in {
                "Ledger Amount",
                "Billed Quantity",
                "Item Rate",
                "Disc%",
                "Item Amount",
                "IGST Rate",
                "CGST Rate",
                "SGST/UTGST Rate",
                "Taxable Value",
            }:
                cell.number_format = "#,##0.00"
    _autosize(sheet)


def import_tally_excel_to_batch(db: Session, batch: Batch, user: User, data: bytes) -> TallyExcelImportResult:
    if batch.status != BatchStatus.DRAFT.value:
        raise InventoryError("Excel import is only available for draft batches")
    if batch.batch_type not in TALLY_EXCEL_IMPORT_BATCH_TYPES:
        raise InventoryError("Excel import is available for sale, issue, and purchase return batches")
    if batch.batch_type == BatchType.SALE.value:
        from app.services.sale_returns import ensure_sale_scan_allowed

        ensure_sale_scan_allowed(db, batch)

    lines = parse_bulk_assignment_xlsx(db, data, user=user, allow_product_create=False)
    total_quantity = sum(line.quantity for line in lines)
    if total_quantity < 1:
        raise InventoryError("Excel file has no importable quantity")
    if total_quantity > MAX_ASSIGNMENT_QUANTITY:
        raise InventoryError(f"Import {MAX_ASSIGNMENT_QUANTITY} items or fewer at a time")

    statuses = fefo_available_statuses(batch.batch_type)
    if not statuses:
        raise InventoryError("No FEFO-ready stock status is configured for this batch")

    picked = _pick_import_serials(db, lines, statuses)
    for line, serials in picked:
        rate = _line_rate(line)
        for serial in serials:
            db.add(BatchItem(batch_id=batch.id, serial_id=serial.id, rate=rate, fefo_picked=True))
            db.add(
                ScanLog(
                    serial_id=serial.id,
                    serial_number_raw=serial.serial_number,
                    user_id=user.id,
                    action=batch.batch_type,
                    batch_id=batch.id,
                    status="EXCEL_IMPORTED",
                    message="Imported from Tally Excel by FEFO",
                )
            )
    db.commit()
    return TallyExcelImportResult(product_lines=len(lines), quantity=total_quantity)


def _pick_import_serials(
    db: Session,
    lines: list[AssignmentLine],
    statuses: set[str],
) -> list[tuple[AssignmentLine, list[Serial]]]:
    selected_ids: set[int] = set()
    picked: list[tuple[AssignmentLine, list[Serial]]] = []
    for line in lines:
        candidates = [
            serial
            for serial in fefo_candidate_serials(
                db,
                line.product.id,
                line.quantity + len(selected_ids),
                statuses=statuses,
            )
            if serial.id not in selected_ids
        ][: line.quantity]
        if len(candidates) < line.quantity:
            available = len(candidates)
            raise InventoryError(
                f"Only {available} FEFO-ready serials are available for {line.product.product_code}"
            )
        selected_ids.update(serial.id for serial in candidates)
        picked.append((line, candidates))
    return picked


def _line_rate(line: AssignmentLine) -> float | None:
    if line.rate is None:
        return None
    if line.rate < Decimal("0"):
        raise InventoryError(f"Rate cannot be negative for {line.product.product_code}")
    return float(line.rate)


def _autosize(sheet) -> None:
    for column in sheet.columns:
        width = max(len(str(cell.value or "")) for cell in column)
        sheet.column_dimensions[column[0].column_letter].width = min(max(width + 2, 12), 42)
