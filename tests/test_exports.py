from datetime import datetime, timezone
from io import BytesIO
from types import SimpleNamespace

from openpyxl import load_workbook
from PIL import Image
from sqlalchemy import select

from app.models import BatchType, InventoryTransaction, Product, ScanLog, SerialStatus, TransactionType, User
from app.services.audit import reconcile_audit_batch
from app.services.exports import (
    audit_reconciliation_xlsx,
    audit_report_pdf,
    barcode_labels_pdf,
    barcode_png,
    label_layout,
    scans_xlsx,
    transactions_xlsx,
)
from app.services.inventory import add_serial_to_batch, apply_batch_statuses, create_batch, generate_serials


def test_scans_xlsx_generates_workbook(db_session):
    user = User(username="admin", password_hash="x", role="admin")
    db_session.add(user)
    db_session.commit()
    scan = ScanLog(serial_number_raw="SG001-000001", user_id=user.id, action="AUDIT", status="SCANNED")
    db_session.add(scan)
    db_session.commit()
    data = scans_xlsx([scan])
    assert data.startswith(b"PK")


def test_scans_xlsx_escapes_formula_like_values(db_session):
    user = User(username="admin", password_hash="x", role="admin")
    db_session.add(user)
    db_session.commit()
    scan = ScanLog(serial_number_raw="=HYPERLINK(\"http://bad\")", user_id=user.id, action="AUDIT", status="SCANNED")
    db_session.add(scan)
    db_session.commit()

    workbook = load_workbook(BytesIO(scans_xlsx([scan])))
    sheet = workbook.active

    assert sheet["D2"].value == "'=HYPERLINK(\"http://bad\")"


def test_scans_xlsx_exports_only_selected_fields(db_session):
    user = User(username="admin", password_hash="x", role="admin")
    db_session.add(user)
    db_session.commit()
    scan = ScanLog(serial_number_raw="SG001-000001", user_id=user.id, action="AUDIT", status="SCANNED")
    db_session.add(scan)
    db_session.commit()

    sheet = load_workbook(BytesIO(scans_xlsx([scan], ["Date", "Serial", "Status"]))).active

    assert [cell.value for cell in sheet[1]] == ["Date", "Serial", "Status"]
    assert [cell.value for cell in sheet[2]][1:] == ["SG001-000001", "SCANNED"]


def test_report_exports_use_dd_mm_yyyy_dates(db_session):
    user = User(username="admin", password_hash="x", role="admin")
    db_session.add(user)
    db_session.commit()
    scan = ScanLog(
        serial_number_raw="SG001-000001",
        user_id=user.id,
        action="AUDIT",
        status="SCANNED",
        created_at=datetime(2026, 6, 28, 9, 0, tzinfo=timezone.utc),
    )
    db_session.add(scan)
    db_session.commit()

    sheet = load_workbook(BytesIO(scans_xlsx([scan]))).active

    assert sheet["A2"].value == "28-06-2026"


def test_audit_reconciliation_xlsx_filters_detail_fields_and_keeps_summary():
    data = audit_reconciliation_xlsx(
        {
            "audit_batch_count": 1,
            "missing": 1,
            "batch_rows": [{"batch_number": "AUD-1", "missing": 1}],
            "product_rows": [{"product_code": "P-1", "missing": 1}],
            "finding_rows": [{"serial_number": "S-1"}],
        },
        ["Audit Batch", "Missing"],
    )

    workbook = load_workbook(BytesIO(data))

    assert workbook.sheetnames == ["Summary", "Audit Batches", "Product Reconciliation", "Serial Findings"]
    assert [cell.value for cell in workbook["Audit Batches"][1]] == ["Audit Batch", "Missing"]
    assert [cell.value for cell in workbook["Product Reconciliation"][1]] == ["Missing"]
    assert [cell.value for cell in workbook["Serial Findings"][1]] == ["Audit Batch"]


def test_barcode_png_is_square_qr_code():
    data = barcode_png("RCV-20260623-0001")
    assert data[:8] == b"\x89PNG\r\n\x1a\n"
    width, height = Image.open(BytesIO(data)).size
    assert width == height
    assert width >= 300


def test_barcode_png_has_scan_quiet_zone():
    data = barcode_png("DIJO-000001")
    image = Image.open(BytesIO(data)).convert("L")
    edge = max(8, image.width // 10)
    top_edge = image.crop((0, 0, image.width, edge))
    left_edge = image.crop((0, 0, edge, image.height))
    bottom_edge = image.crop((0, image.height - edge, image.width, image.height))
    right_edge = image.crop((image.width - edge, 0, image.width, image.height))

    assert top_edge.getextrema()[0] > 245
    assert left_edge.getextrema()[0] > 245
    assert bottom_edge.getextrema()[0] > 245
    assert right_edge.getextrema()[0] > 245


def test_label_layout_defaults_to_48_5_by_25_4_sheet_grid():
    assert label_layout() == (11, 4)


def _make_product(db_session, code: str) -> Product:
    product = Product(
        product_code=code,
        product_name="Masala",
        hsn="0910",
        gst_rate=5,
        unit="Pcs",
        default_rate=100,
        tally_stock_item_name="Masala",
    )
    db_session.add(product)
    db_session.commit()
    return product


def test_barcode_labels_pdf_generates_pdf(db_session):
    product = _make_product(db_session, "SG050")
    serial = generate_serials(db_session, product, 1)[0]
    data = barcode_labels_pdf([serial])
    assert data.startswith(b"%PDF")


def test_barcode_labels_pdf_accepts_custom_page_grid(db_session):
    product = _make_product(db_session, "SG053")
    serials = generate_serials(db_session, product, 5)
    data = barcode_labels_pdf(serials, rows_per_page=2, columns_per_page=2)
    assert data.startswith(b"%PDF")


def test_barcode_labels_pdf_renders_full_qr_label_sheet():
    serials = [SimpleNamespace(serial_number=f"DIJO-{index:06d}") for index in range(44)]
    data = barcode_labels_pdf(serials)
    assert data.startswith(b"%PDF")


def test_transactions_xlsx_includes_edit_log_actor_columns(db_session):
    sales_user = User(username="sales", password_hash="x", role="sales")
    product = _make_product(db_session, "SG052")
    db_session.add(sales_user)
    db_session.commit()
    serial = generate_serials(db_session, product, 1, initial_status=SerialStatus.IN_STOCK)[0]
    batch = create_batch(db_session, sales_user, BatchType.SALE, "Customer", "")
    add_serial_to_batch(db_session, batch, sales_user, serial.serial_number)
    apply_batch_statuses(db_session, batch, sales_user)
    txn = db_session.scalar(select(InventoryTransaction).where(InventoryTransaction.transaction_type == TransactionType.SALE.value))

    data = transactions_xlsx([txn])
    workbook = load_workbook(BytesIO(data))
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    values = [cell.value for cell in sheet[2]]

    assert "Invoice Created By" in headers
    assert "Barcode Sold By" in headers
    assert "Product Audited By" in headers
    assert values[headers.index("Invoice Created By")] == "sales"
    assert values[headers.index("Barcode Sold By")] == "sales"
    assert values[headers.index("Product Audited By")] is None


def test_audit_report_pdf_generates_pdf(db_session):
    auditor = User(username="auditor", password_hash="x", role="auditor")
    product = _make_product(db_session, "SG051")
    db_session.add(auditor)
    db_session.commit()
    serial = generate_serials(db_session, product, 1, initial_status=SerialStatus.IN_STOCK)[0]
    batch = create_batch(db_session, auditor, BatchType.AUDIT, "Rack A", "")
    add_serial_to_batch(db_session, batch, auditor, serial.serial_number)
    reconcile_audit_batch(db_session, batch)
    data = audit_report_pdf(batch)
    assert data.startswith(b"%PDF")
