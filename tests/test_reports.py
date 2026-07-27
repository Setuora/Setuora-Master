import json
from datetime import datetime, timedelta, timezone
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool
from starlette.requests import Request

from app.auth import SESSION_COOKIE
from app.database import Base, get_db
from app.main import app
from app.models import AuditAssignment, AuditAssignmentItem, AuditFinding, Batch, BatchItem, BatchStatus, BatchType, InventoryTransaction, Product, ScanLog, Serial, SerialStatus, StorageLocation, TransactionType, User, WarehouseLevel
from app.routers.dashboard import dashboard, dashboard_data
from app.routers.reports import (
    audit_reconciliation_excel,
    director_audit_batch_detail,
    director_report_live,
    missing_stock_excel,
    missing_stock_report,
    reports as reports_route,
)
from app.security import create_session_token
from app.services.access_control import save_role_access_config
from app.services.expiry import today
from app.services.director_reports import (
    director_product_stock_rows,
    director_warehouse_filter_options,
    director_warehouse_stock_rows,
)


def test_reports_page_renders_scan_and_transaction_rows():
    engine = create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    Session = sessionmaker(bind=engine)

    with Session() as db:
        user = User(id=1, username="admin", password_hash="x", role="admin", active=True)
        product = Product(
            product_code="SG100",
            product_name="Masala",
            hsn="0910",
            gst_rate=5,
            unit="Pcs",
            default_rate=100,
            tally_stock_item_name="Masala",
        )
        db.add_all([user, product])
        db.flush()
        audit_product = Product(
            product_code="AUD100",
            product_name="Audit Progress Masala",
            hsn="0910",
            gst_rate=5,
            unit="Pcs",
            default_rate=100,
            tally_stock_item_name="Audit Progress Masala",
        )
        db.add(audit_product)
        db.flush()
        first_expected = Serial(
            serial_number="AUD100-000001",
            product_id=audit_product.id,
            status=SerialStatus.IN_STOCK.value,
        )
        second_expected = Serial(
            serial_number="AUD100-000002",
            product_id=audit_product.id,
            status=SerialStatus.IN_STOCK.value,
        )
        db.add_all([first_expected, second_expected])
        db.flush()
        now = datetime.now(timezone.utc)
        audit_assignment = AuditAssignment(
            product_id=audit_product.id,
            auditor_id=user.id,
            assigned_by_id=user.id,
            starts_at=now - timedelta(hours=1),
            ends_at=now + timedelta(hours=1),
        )
        db.add(audit_assignment)
        db.flush()
        db.add_all(
            [
                AuditAssignmentItem(assignment_id=audit_assignment.id, serial_id=first_expected.id),
                AuditAssignmentItem(assignment_id=audit_assignment.id, serial_id=second_expected.id),
            ]
        )
        audit_batch = Batch(
            batch_number="AUD-PROG-001",
            batch_type=BatchType.AUDIT.value,
            user_id=user.id,
            audit_assignment_id=audit_assignment.id,
            status=BatchStatus.SUBMITTED.value,
            submitted_at=now,
        )
        db.add(audit_batch)
        db.flush()
        db.add(BatchItem(batch_id=audit_batch.id, serial_id=first_expected.id, created_at=now))
        batch = Batch(
            batch_number="BATCH-001",
            batch_type=BatchType.SALE.value,
            user_id=user.id,
            status=BatchStatus.PENDING_SYNC.value,
        )
        db.add(batch)
        db.flush()
        db.add_all(
            [
                InventoryTransaction(
                    transaction_type=TransactionType.SALE.value,
                    product_id=product.id,
                    batch_id=batch.id,
                    user_id=user.id,
                    serial_number="SG100-000001",
                    status_from="IN_STOCK",
                    status_to="SOLD",
                    reference_number=batch.batch_number,
                    tally_reference="TALLY-001",
                ),
                ScanLog(
                    serial_number_raw="SG100-000001",
                    user_id=user.id,
                    action=BatchType.SALE.value,
                    batch_id=batch.id,
                    status="SCANNED",
                ),
            ]
        )
        db.commit()

    def override_get_db():
        db = Session()
        try:
            yield db
        finally:
            db.close()

    app.dependency_overrides[get_db] = override_get_db
    try:
        client = TestClient(app, follow_redirects=False, headers={"Origin": "http://testserver"})
        response = client.get("/reports", cookies={SESSION_COOKIE: create_session_token(1)})
    finally:
        app.dependency_overrides.clear()
        engine.dispose()

    assert response.status_code == 200
    assert "Transaction mix" in response.text
    assert "Scan outcomes" in response.text
    assert "SG100-000001" in response.text
    assert "TALLY-001" in response.text
    assert "BATCH-001" in response.text
    assert "Audit assignments" in response.text
    assert "Expected" in response.text
    assert "Found" in response.text
    assert "Remaining" in response.text
    audit_section_start = response.text.index("<h2>Audit assignments</h2>")
    row_start = response.text.index("Audit Progress Masala", audit_section_start)
    assignment_row = response.text[row_start:response.text.index("</tr>", row_start)]
    assert "<td>2</td>" in assignment_row
    assert assignment_row.count("<td>1</td>") >= 2


def test_reports_product_dropdown_filters_product_specific_rows_and_exports():
    engine = create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    Session = sessionmaker(bind=engine)

    with Session() as db:
        user = User(id=1, username="admin", password_hash="x", role="admin", active=True)
        selected = Product(
            product_code="SEL100",
            product_name="Selected Masala",
            hsn="0910",
            gst_rate=5,
            unit="Pcs",
            default_rate=100,
            tally_stock_item_name="Selected Masala",
        )
        other = Product(
            product_code="OTH200",
            product_name="Other Pepper",
            hsn="0910",
            gst_rate=5,
            unit="Pcs",
            default_rate=120,
            tally_stock_item_name="Other Pepper",
        )
        db.add_all([user, selected, other])
        db.flush()
        selected_serial = Serial(
            serial_number="SEL100-000001",
            product_id=selected.id,
            status=SerialStatus.IN_STOCK.value,
            product_batch_number="SEL-EXP-BATCH",
            expiry_date=today() + timedelta(days=20),
        )
        other_serial = Serial(
            serial_number="OTH200-000001",
            product_id=other.id,
            status=SerialStatus.IN_STOCK.value,
            product_batch_number="OTH-EXP-BATCH",
            expiry_date=today() + timedelta(days=20),
        )
        db.add_all([selected_serial, other_serial])
        db.flush()
        selected_batch = Batch(
            batch_number="SEL-BATCH-001",
            batch_type=BatchType.SALE.value,
            user_id=user.id,
            status=BatchStatus.SUBMITTED.value,
        )
        other_batch = Batch(
            batch_number="OTH-BATCH-001",
            batch_type=BatchType.SALE.value,
            user_id=user.id,
            status=BatchStatus.SUBMITTED.value,
        )
        db.add_all([selected_batch, other_batch])
        db.flush()
        db.add_all(
            [
                InventoryTransaction(
                    transaction_type=TransactionType.SALE.value,
                    product_id=selected.id,
                    serial_id=selected_serial.id,
                    batch_id=selected_batch.id,
                    user_id=user.id,
                    serial_number=selected_serial.serial_number,
                    status_from=SerialStatus.IN_STOCK.value,
                    status_to=SerialStatus.SOLD.value,
                    reference_number=selected_batch.batch_number,
                    tally_reference="SEL-TALLY-001",
                ),
                InventoryTransaction(
                    transaction_type=TransactionType.SALE.value,
                    product_id=other.id,
                    serial_id=other_serial.id,
                    batch_id=other_batch.id,
                    user_id=user.id,
                    serial_number=other_serial.serial_number,
                    status_from=SerialStatus.IN_STOCK.value,
                    status_to=SerialStatus.SOLD.value,
                    reference_number=other_batch.batch_number,
                    tally_reference="OTH-TALLY-001",
                ),
                ScanLog(
                    serial_id=selected_serial.id,
                    serial_number_raw=selected_serial.serial_number,
                    user_id=user.id,
                    action=BatchType.SALE.value,
                    status="SCANNED",
                ),
                ScanLog(
                    serial_id=other_serial.id,
                    serial_number_raw=other_serial.serial_number,
                    user_id=user.id,
                    action=BatchType.SALE.value,
                    status="SCANNED",
                ),
                AuditFinding(
                    batch_id=selected_batch.id,
                    serial_id=selected_serial.id,
                    serial_number=selected_serial.serial_number,
                    product_code=selected.product_code,
                    product_name=selected.product_name,
                    finding_type="MISSING",
                    expected_status=SerialStatus.IN_STOCK.value,
                ),
                AuditFinding(
                    batch_id=other_batch.id,
                    serial_id=other_serial.id,
                    serial_number=other_serial.serial_number,
                    product_code=other.product_code,
                    product_name=other.product_name,
                    finding_type="MISSING",
                    expected_status=SerialStatus.IN_STOCK.value,
                ),
            ]
        )
        selected_id = selected.id
        db.commit()

    def override_get_db():
        db = Session()
        try:
            yield db
        finally:
            db.close()

    app.dependency_overrides[get_db] = override_get_db
    try:
        client = TestClient(app, follow_redirects=False, headers={"Origin": "http://testserver"})
        response = client.get(f"/reports?product_id={selected_id}", cookies={SESSION_COOKIE: create_session_token(1)})
        all_products_response = client.get(
            "/reports?action=&product_id=&start=&end=",
            cookies={SESSION_COOKIE: create_session_token(1)},
        )
        invalid_product_response = client.get(
            "/reports?product_id=not-a-product",
            cookies={SESSION_COOKIE: create_session_token(1)},
        )
        export = client.get(
            f"/reports/transactions.xlsx?product_id={selected_id}",
            cookies={SESSION_COOKIE: create_session_token(1)},
        )
    finally:
        app.dependency_overrides.clear()
        engine.dispose()

    assert response.status_code == 200
    assert f'<option value="{selected_id}" selected>SEL100 - Selected Masala</option>' in response.text
    assert "SEL100-000001" in response.text
    assert "SEL-TALLY-001" in response.text
    assert "SEL-EXP-BATCH" in response.text
    assert "OTH200-000001" not in response.text
    assert "OTH-TALLY-001" not in response.text
    assert "OTH-EXP-BATCH" not in response.text
    assert "product_id" in response.text

    assert all_products_response.status_code == 200
    assert "SEL100-000001" in all_products_response.text
    assert "OTH200-000001" in all_products_response.text
    assert invalid_product_response.status_code == 400
    assert invalid_product_response.json()["detail"] == "Invalid product filter"

    workbook = load_workbook(BytesIO(export.content))
    sheet = workbook["Transactions"]
    rows = list(sheet.iter_rows(values_only=True))
    assert rows[1][3] == "SEL100-000001"
    assert all("OTH200-000001" not in row for row in rows if row)


def test_reports_page_includes_filterable_missing_stock():
    engine = create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    Session = sessionmaker(bind=engine)

    with Session() as db:
        user = User(id=1, username="auditor", password_hash="x", role="admin", active=True)
        product = Product(
            product_code="MISS100",
            product_name="Missing masala",
            hsn="0910",
            gst_rate=5,
            unit="Pcs",
            default_rate=100,
            tally_stock_item_name="Missing masala",
        )
        db.add_all([user, product])
        db.flush()
        serial = Serial(
            serial_number="MISS100-000001",
            product_id=product.id,
            status=SerialStatus.IN_STOCK.value,
            product_batch_number="LOT-MISS-01",
            warehouse="Main warehouse",
            mfg_date=today() - timedelta(days=30),
            expiry_date=today() + timedelta(days=120),
        )
        resolved_serial = Serial(
            serial_number="MISS100-000002",
            product_id=product.id,
            status=SerialStatus.IN_STOCK.value,
            product_batch_number="LOT-MISS-02",
            warehouse="Main warehouse",
        )
        batch = Batch(
            batch_number="AUD-001",
            batch_type=BatchType.AUDIT.value,
            user_id=user.id,
            status=BatchStatus.SUBMITTED.value,
            submitted_at=datetime(2026, 6, 28, 9, 0, tzinfo=timezone.utc),
        )
        resolved_batch = Batch(
            batch_number="AUD-002",
            batch_type=BatchType.AUDIT.value,
            user_id=user.id,
            status=BatchStatus.SUBMITTED.value,
            submitted_at=datetime(2026, 6, 29, 9, 0, tzinfo=timezone.utc),
        )
        db.add_all([serial, resolved_serial, batch, resolved_batch])
        db.flush()
        db.add_all(
            [
                AuditFinding(
                    batch_id=batch.id,
                    serial_id=serial.id,
                    serial_number=serial.serial_number,
                    product_code=product.product_code,
                    product_name=product.product_name,
                    finding_type="MISSING",
                    expected_status=SerialStatus.IN_STOCK.value,
                    created_at=datetime(2026, 6, 28, 9, 5, tzinfo=timezone.utc),
                ),
                AuditFinding(
                    batch_id=batch.id,
                    serial_id=resolved_serial.id,
                    serial_number=resolved_serial.serial_number,
                    product_code=product.product_code,
                    product_name=product.product_name,
                    finding_type="MISSING",
                    expected_status=SerialStatus.IN_STOCK.value,
                    created_at=datetime(2026, 6, 28, 9, 5, tzinfo=timezone.utc),
                ),
                AuditFinding(
                    batch_id=resolved_batch.id,
                    serial_id=resolved_serial.id,
                    serial_number=resolved_serial.serial_number,
                    product_code=product.product_code,
                    product_name=product.product_name,
                    finding_type="VERIFIED",
                    expected_status=SerialStatus.IN_STOCK.value,
                    scanned_status=SerialStatus.IN_STOCK.value,
                    created_at=datetime(2026, 6, 29, 9, 5, tzinfo=timezone.utc),
                ),
            ]
        )
        db.commit()

    def signed_request(path: str, query_string: bytes = b"") -> Request:
        return Request(
            {
                "type": "http",
                "method": "GET",
                "path": path,
                "headers": [(b"cookie", f"{SESSION_COOKIE}={create_session_token(1)}".encode())],
                "query_string": query_string,
                "server": ("testserver", 80),
                "scheme": "http",
            }
        )

    with Session() as db:
        response = reports_route(
            signed_request("/reports", b"action=MISSING"),
            action="MISSING",
            db=db,
        )
        detail_response = missing_stock_report(
            signed_request("/reports/missing-stock"),
            db=db,
        )
        xlsx_response = missing_stock_excel(
            signed_request("/reports/missing-stock.xlsx", b"q=MISS100"),
            q="MISS100",
            db=db,
        )
    engine.dispose()

    response_text = response.body.decode()
    detail_text = detail_response.body.decode()

    assert response.status_code == 200
    assert '<option value="MISSING" selected>MISSING</option>' in response_text
    assert "<h2>Missing stock</h2>" in response_text
    assert "MISS100-000001" in response_text
    assert "28-06-2026" in response_text
    assert "MISS100-000002" not in response_text
    assert "Missing masala" in response_text
    assert "AUD-001" in response_text
    assert "AUD-002" not in response_text
    assert "Missing stock CSV" not in response_text
    assert "Missing stock Excel" in response_text
    assert "data-xlsx-export" in response_text
    assert "Customize missing stock export" in response_text
    assert "Serial or reference" not in response_text
    assert 'data-xlsx-parameters="q|product_id|start|end"' not in response_text

    assert detail_response.status_code == 200
    assert "<h1>Missing stock report</h1>" in detail_text
    assert "<h2>Missing stock details</h2>" in detail_text
    assert "LOT-MISS-01" in detail_text
    assert "Main warehouse" in detail_text
    assert "MISS100-000001" in detail_text
    assert "28-06-2026" in detail_text
    assert "MISS100-000002" not in detail_text
    assert 'href="/reports/missing-stock"' in detail_text
    assert ">Overview</a>" in detail_text
    assert "Serial or audit batch" not in detail_text
    assert 'data-xlsx-parameters="q|product_id|start|end"' not in detail_text

    assert xlsx_response.status_code == 200
    assert xlsx_response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert xlsx_response.body.startswith(b"PK")


def test_directors_role_gets_report_only_summary_and_audit_detail():
    engine = create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    Session = sessionmaker(bind=engine)

    audit_at = datetime(2026, 6, 28, 10, 30, tzinfo=timezone.utc)
    old_at = datetime.now(timezone.utc) - timedelta(days=140)
    with Session() as db:
        auditor = User(id=1, username="auditor", password_hash="x", role="admin", active=True)
        director = User(id=2, username="director", password_hash="x", role="directors", active=True)
        missing_product = Product(
            product_code="DIR-MISS",
            product_name="Director missing product",
            hsn="0910",
            gst_rate=5,
            unit="Pcs",
            default_rate=100,
            tally_stock_item_name="Director missing product",
        )
        risk_product = Product(
            product_code="DIR-RISK",
            product_name="Director expiry risk product",
            hsn="0910",
            gst_rate=5,
            unit="Pcs",
            default_rate=80,
            tally_stock_item_name="Director expiry risk product",
        )
        dead_product = Product(
            product_code="DIR-DEAD",
            product_name="Director dead stock product",
            hsn="0910",
            gst_rate=5,
            unit="Pcs",
            default_rate=60,
            tally_stock_item_name="Director dead stock product",
            created_at=old_at,
        )
        db.add_all([auditor, director, missing_product, risk_product, dead_product])
        db.flush()
        missing_serial = Serial(
            serial_number="DIR-MISS-001",
            product_id=missing_product.id,
            status=SerialStatus.IN_STOCK.value,
        )
        extra_serial = Serial(
            serial_number="DIR-MISS-EXTRA",
            product_id=missing_product.id,
            status=SerialStatus.SOLD.value,
        )
        risk_serial = Serial(
            serial_number="DIR-RISK-001",
            product_id=risk_product.id,
            status=SerialStatus.IN_STOCK.value,
            product_batch_number="RISK-B1",
            expiry_date=today() + timedelta(days=20),
        )
        dead_serial = Serial(
            serial_number="DIR-DEAD-001",
            product_id=dead_product.id,
            status=SerialStatus.IN_STOCK.value,
        )
        db.add_all([missing_serial, extra_serial, risk_serial, dead_serial])
        db.flush()
        batch = Batch(
            batch_number="AUD-DIR-001",
            batch_type=BatchType.AUDIT.value,
            user_id=auditor.id,
            status=BatchStatus.SUBMITTED.value,
            submitted_at=audit_at,
        )
        db.add(batch)
        db.flush()
        batch_id = batch.id
        db.add_all(
            [
                AuditFinding(
                    batch_id=batch.id,
                    serial_id=missing_serial.id,
                    serial_number=missing_serial.serial_number,
                    product_code=missing_product.product_code,
                    product_name=missing_product.product_name,
                    finding_type="MISSING",
                    expected_status=SerialStatus.IN_STOCK.value,
                ),
                AuditFinding(
                    batch_id=batch.id,
                    serial_id=extra_serial.id,
                    serial_number=extra_serial.serial_number,
                    product_code=missing_product.product_code,
                    product_name=missing_product.product_name,
                    finding_type="EXTRA",
                    expected_status=SerialStatus.IN_STOCK.value,
                    scanned_status=SerialStatus.SOLD.value,
                ),
            ]
        )
        db.commit()

    def signed_request(path: str) -> Request:
        return Request(
            {
                "type": "http",
                "method": "GET",
                "path": path,
                "headers": [(b"cookie", f"{SESSION_COOKIE}={create_session_token(2)}".encode())],
                "query_string": b"",
                "server": ("testserver", 80),
                "scheme": "http",
            }
        )

    with Session() as db:
        report_response = reports_route(signed_request("/reports"), db=db)
        detail_response = director_audit_batch_detail(
            signed_request(f"/reports/audit-batches/{batch_id}"),
            batch_id,
            db=db,
        )
        report_text = report_response.body.decode()
        detail_text = detail_response.body.decode()
    engine.dispose()

    assert report_response.status_code == 200
    assert report_response.template.name == "director_reports.html"
    assert "Directors Report" in report_text
    assert "Reports only" in report_text
    assert "AUD-DIR-001" in report_text
    assert "Missing in last audit" in report_text
    assert "Director expiry risk product" in report_text
    assert "Director dead stock product" in report_text
    assert "Live data" in report_text
    assert "Product stock" in report_text
    assert "Warehouse level totals" in report_text
    assert "Audit reconciliation Excel" in report_text
    assert "Audit reconciliation" in report_text
    assert "Transactions Excel" not in report_text
    assert "<h2>Transactions</h2>" not in report_text
    assert 'href="/reports"' in report_text
    assert ">Dashboard</a>" not in report_text
    assert ">Serials</a>" not in report_text

    assert detail_response.status_code == 200
    assert detail_response.template.name == "director_audit_batch.html"
    assert "Product-wise missing and extra" in detail_text
    assert "Director missing product" in detail_text
    assert "DIR-MISS-001" in detail_text
    assert "DIR-MISS-EXTRA" in detail_text


def test_director_report_groups_current_stock_by_warehouse_and_prioritizes_it():
    engine = create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    Session = sessionmaker(bind=engine)

    with Session() as db:
        director = User(id=1, username="director", password_hash="x", role="directors", active=True)
        product = Product(
            product_code="WH100",
            product_name="Warehouse stock product",
            hsn="0910",
            gst_rate=5,
            unit="Pcs",
            default_rate=100,
            tally_stock_item_name="Warehouse stock product",
        )
        main_location = StorageLocation(
            code="MAIN-A-1-1-1-1",
            warehouse="Main warehouse",
            zone="A",
            section="1",
            rack="1",
            shelf="1",
            bin="1",
        )
        db.add_all([director, product, main_location])
        db.flush()
        db.add_all(
            [
                Serial(
                    serial_number="WH100-000001",
                    product_id=product.id,
                    status=SerialStatus.IN_STOCK.value,
                    warehouse="Old warehouse name",
                    location_id=main_location.id,
                    expiry_date=today() + timedelta(days=20),
                ),
                Serial(
                    serial_number="WH100-000002",
                    product_id=product.id,
                    status=SerialStatus.IN_STOCK.value,
                    warehouse="Main warehouse",
                ),
                Serial(
                    serial_number="WH100-000003",
                    product_id=product.id,
                    status=SerialStatus.IN_STOCK.value,
                    warehouse="Branch warehouse",
                    warehouse_level=WarehouseLevel.C_AND_F.value,
                ),
            ]
        )
        db.commit()

    def signed_request() -> Request:
        return Request(
            {
                "type": "http",
                "method": "GET",
                "path": "/reports",
                "headers": [(b"cookie", f"{SESSION_COOKIE}={create_session_token(1)}".encode())],
                "query_string": b"",
                "server": ("testserver", 80),
                "scheme": "http",
            }
        )

    with Session() as db:
        response = reports_route(signed_request(), db=db)
        report_text = response.body.decode()
        live_response = director_report_live(signed_request(), db=db)
        live_data = json.loads(live_response.body)
        warehouse_rows = director_warehouse_stock_rows(db)
        filtered_response = reports_route(
            signed_request(),
            product_q="Warehouse stock",
            warehouse_q=WarehouseLevel.C_AND_F.value,
            db=db,
        )
        filtered_text = filtered_response.body.decode()
        filtered_live_response = director_report_live(
            signed_request(),
            product_q="Warehouse stock",
            warehouse_q=WarehouseLevel.C_AND_F.value,
            db=db,
        )
        filtered_live_data = json.loads(filtered_live_response.body)
    engine.dispose()

    assert response.status_code == 200
    assert live_response.status_code == 200
    assert 'data-live-url="/reports/live"' in report_text
    assert report_text.index("<h2>Product stock</h2>") < report_text.index("<h2>Audit reconciliation</h2>")
    assert report_text.index("<h2>Warehouse level totals</h2>") < report_text.index("<h2>Audit reconciliation</h2>")
    warehouse_table_rows = report_text.split(
        "data-director-live-warehouse-rows>", 1
    )[1].split("</tbody>", 1)[0]
    assert WarehouseLevel.COMPANY_WAREHOUSE.value in warehouse_table_rows
    assert "C&amp;F" in warehouse_table_rows
    assert "Old warehouse name" not in warehouse_table_rows
    assert live_data["director_metrics"]["total_stock"] == 3
    assert live_data["director_metrics"]["total_products"] == 1
    assert WarehouseLevel.COMPANY_WAREHOUSE.value in live_data["warehouse_rows_html"]
    assert "C&amp;F" in live_data["warehouse_rows_html"]
    assert "Products" in report_text
    assert "Locations" in report_text
    assert "Unlocated" in report_text
    assert "Expiry ≤30 days" in report_text
    assert "Nearest expiry" in report_text
    company_warehouse = next(
        row
        for row in warehouse_rows
        if row["warehouse_level"] == WarehouseLevel.COMPANY_WAREHOUSE.value
    )
    assert company_warehouse["stock"] == 2
    assert company_warehouse["warehouses"] == 1
    assert company_warehouse["products"] == 1
    assert company_warehouse["locations"] == 1
    assert company_warehouse["unlocated"] == 1
    assert company_warehouse["expiring_soon"] == 1
    assert company_warehouse["nearest_expiry"] == today() + timedelta(days=20)
    assert 'data-director-stock-filter' in report_text
    assert 'aria-label="Filter product stock by product"' in report_text
    assert 'aria-label="Filter warehouse totals by warehouse level"' in report_text
    assert '<option value="WH100"' in report_text
    assert '<option value="C&amp;F"' in report_text
    assert 'data-live-url="/reports/live?product_q=Warehouse+stock&amp;warehouse_q=C%26F"' in filtered_text
    filtered_warehouse_rows = filtered_text.split('data-director-live-warehouse-rows>', 1)[1].split("</tbody>", 1)[0]
    assert WarehouseLevel.COMPANY_WAREHOUSE.value not in filtered_warehouse_rows
    assert "C&amp;F" in filtered_warehouse_rows
    assert WarehouseLevel.COMPANY_WAREHOUSE.value not in filtered_live_data["warehouse_rows_html"]
    assert "C&amp;F" in filtered_live_data["warehouse_rows_html"]


def test_director_warehouse_filter_uses_levels_instead_of_warehouse_names():
    engine = create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    Session = sessionmaker(bind=engine)

    with Session() as db:
        product = Product(
            product_code="WH-LEVEL",
            product_name="Warehouse level guard",
            hsn="0910",
            gst_rate=5,
            unit="Pcs",
            default_rate=100,
            tally_stock_item_name="Warehouse level guard",
        )
        actual_location = StorageLocation(
            code="BENGALURU-A-1-1-1-1",
            warehouse="Bengaluru warehouse",
            warehouse_level=WarehouseLevel.COMPANY_WAREHOUSE.value,
            zone="A",
            section="1",
            rack="1",
            shelf="1",
            bin="1",
        )
        db.add_all([product, actual_location])
        db.flush()
        db.add_all(
            [
                Serial(
                    serial_number="WH-LEVEL-000001",
                    product_id=product.id,
                    status=SerialStatus.IN_STOCK.value,
                    warehouse=WarehouseLevel.COMPANY_WAREHOUSE.value,
                    warehouse_level=WarehouseLevel.COMPANY_WAREHOUSE.value,
                ),
                Serial(
                    serial_number="WH-LEVEL-000002",
                    product_id=product.id,
                    status=SerialStatus.IN_STOCK.value,
                    warehouse=WarehouseLevel.C_AND_F.value,
                    warehouse_level=WarehouseLevel.C_AND_F.value,
                ),
                Serial(
                    serial_number="WH-LEVEL-000003",
                    product_id=product.id,
                    status=SerialStatus.IN_STOCK.value,
                    warehouse=WarehouseLevel.COMPANY_WAREHOUSE.value,
                    warehouse_level=WarehouseLevel.COMPANY_WAREHOUSE.value,
                    location_id=actual_location.id,
                ),
            ]
        )
        db.commit()

        options = director_warehouse_filter_options(db)
        rows = director_warehouse_stock_rows(db)

    engine.dispose()

    assert options == [
        WarehouseLevel.COMPANY_WAREHOUSE.value,
        WarehouseLevel.C_AND_F.value,
        WarehouseLevel.MASTER_FRANCHISE.value,
        WarehouseLevel.TALUK_FRANCHISE.value,
        WarehouseLevel.HOME_FRANCHISE.value,
    ]
    assert {row["warehouse_level"]: row["stock"] for row in rows} == {
        WarehouseLevel.COMPANY_WAREHOUSE.value: 2,
        WarehouseLevel.C_AND_F.value: 1,
    }
    assert "Bengaluru warehouse" not in options


def test_admin_and_director_stock_summary_is_cumulative_and_product_searchable():
    engine = create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    Session = sessionmaker(bind=engine)
    audit_at = datetime(2026, 7, 12, 10, 30, tzinfo=timezone.utc)

    with Session() as db:
        admin = User(id=1, username="admin", password_hash="x", role="admin", active=True)
        director = User(id=2, username="director", password_hash="x", role="directors", active=True)
        primary = Product(
            product_code="SUM-A",
            product_name="Summary primary product",
            hsn="0910",
            gst_rate=5,
            unit="Pcs",
            default_rate=100,
            tally_stock_item_name="Summary primary product",
        )
        other = Product(
            product_code="SUM-B",
            product_name="Summary other product",
            hsn="0910",
            gst_rate=5,
            unit="Pcs",
            default_rate=80,
            tally_stock_item_name="Summary other product",
        )
        db.add_all([admin, director, primary, other])
        db.flush()
        stock_serial = Serial(
            serial_number="SUM-A-000001",
            product_id=primary.id,
            status=SerialStatus.IN_STOCK.value,
        )
        sold_serial = Serial(
            serial_number="SUM-A-000002",
            product_id=primary.id,
            status=SerialStatus.SOLD.value,
        )
        other_serial = Serial(
            serial_number="SUM-B-000001",
            product_id=other.id,
            status=SerialStatus.RETURNED.value,
        )
        db.add_all([stock_serial, sold_serial, other_serial])
        db.flush()
        audit_batch = Batch(
            batch_number="AUD-SUM-001",
            batch_type=BatchType.AUDIT.value,
            user_id=admin.id,
            status=BatchStatus.SUBMITTED.value,
            submitted_at=audit_at,
        )
        db.add(audit_batch)
        db.flush()
        db.add_all(
            [
                InventoryTransaction(
                    transaction_type=TransactionType.PURCHASE.value,
                    product_id=primary.id,
                    serial_id=stock_serial.id,
                    user_id=admin.id,
                ),
                InventoryTransaction(
                    transaction_type=TransactionType.PURCHASE.value,
                    product_id=primary.id,
                    serial_id=sold_serial.id,
                    user_id=admin.id,
                ),
                InventoryTransaction(
                    transaction_type=TransactionType.SALE.value,
                    product_id=primary.id,
                    serial_id=sold_serial.id,
                    user_id=admin.id,
                ),
                AuditFinding(
                    batch_id=audit_batch.id,
                    serial_id=stock_serial.id,
                    serial_number=stock_serial.serial_number,
                    product_code=primary.product_code,
                    product_name=primary.product_name,
                    finding_type="VERIFIED",
                    expected_status=SerialStatus.IN_STOCK.value,
                    scanned_status=SerialStatus.IN_STOCK.value,
                ),
                AuditFinding(
                    batch_id=audit_batch.id,
                    serial_id=sold_serial.id,
                    serial_number=sold_serial.serial_number,
                    product_code=primary.product_code,
                    product_name=primary.product_name,
                    finding_type="EXTRA",
                    expected_status=SerialStatus.IN_STOCK.value,
                    scanned_status=SerialStatus.SOLD.value,
                ),
                AuditFinding(
                    batch_id=audit_batch.id,
                    serial_id=other_serial.id,
                    serial_number=other_serial.serial_number,
                    product_code=other.product_code,
                    product_name=other.product_name,
                    finding_type="MISSING",
                    expected_status=SerialStatus.IN_STOCK.value,
                ),
            ]
        )
        db.commit()

    def signed_request(path: str, user_id: int) -> Request:
        return Request(
            {
                "type": "http",
                "method": "GET",
                "path": path,
                "headers": [
                    (b"cookie", f"{SESSION_COOKIE}={create_session_token(user_id)}".encode())
                ],
                "query_string": b"",
                "server": ("testserver", 80),
                "scheme": "http",
            }
        )

    with Session() as db:
        rows = director_product_stock_rows(db)
        filtered_rows = director_product_stock_rows(db, "SUM-A")
        admin_response = dashboard(signed_request("/", 1), db=db)
        admin_filtered = dashboard(signed_request("/", 1), product_q="SUM-A", db=db)
        admin_live = dashboard_data(
            signed_request("/dashboard/data", 1), product_q="SUM-A", db=db
        )
        director_response = reports_route(signed_request("/reports", 2), db=db)
        director_filtered = reports_route(
            signed_request("/reports", 2), product_q="SUM-A", db=db
        )
        director_live = director_report_live(
            signed_request("/reports/live", 2),
            db=db,
        )
    engine.dispose()

    primary_row = next(row for row in rows if row["product_code"] == "SUM-A")
    assert primary_row["stock"] == 1
    assert primary_row["purchased"] == 2
    assert primary_row["sold"] == 1
    assert primary_row["last_audit_at"] == audit_at.replace(tzinfo=None)
    assert primary_row["last_audited_quantity"] == 2
    assert primary_row["last_audit_missing"] == 0
    assert primary_row["last_audit_extra"] == 1
    other_row = next(row for row in rows if row["product_code"] == "SUM-B")
    assert other_row["last_audit_at"] == audit_at.replace(tzinfo=None)
    assert other_row["last_audited_quantity"] == 0
    assert other_row["last_audit_missing"] == 1
    assert other_row["last_audit_extra"] == 0
    assert [row["product_code"] for row in filtered_rows] == ["SUM-A"]

    for response in (admin_response, director_response):
        text = response.body.decode()
        assert "Current stock" in text
        assert "Purchased (cumulative)" in text
        assert "Sold (cumulative)" in text
        assert "Last audit" in text
        assert "Last audited qty" in text
        assert "Summary primary product" in text
        assert "Summary other product" in text

    director_text = director_response.body.decode()
    assert "Last audit missing" in director_text
    assert "Last audit extra" in director_text
    assert "Nothing scanned" in director_text
    assert "Last audit missing" not in admin_response.body.decode()

    director_live_rows = json.loads(director_live.body)["product_rows_html"]
    assert "Nothing scanned" in director_live_rows
    assert 'class="status missing">1</span>' in director_live_rows

    admin_filtered_rows = admin_filtered.body.decode().split(
        "data-dashboard-product-stock-rows>", 1
    )[1].split("</tbody>", 1)[0]
    director_filtered_rows = director_filtered.body.decode().split(
        "data-director-live-product-rows>", 1
    )[1].split("</tbody>", 1)[0]
    live_rows = json.loads(admin_live.body)["product_stock_rows_html"]
    for html in (admin_filtered_rows, director_filtered_rows, live_rows):
        assert "Summary primary product" in html
        assert "Summary other product" not in html


def test_audit_reconciliation_xlsx_combines_audit_batches_for_admin_and_director():
    engine = create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    Session = sessionmaker(bind=engine)

    with Session() as db:
        admin = User(id=1, username="admin", password_hash="x", role="admin", active=True)
        director = User(id=2, username="director", password_hash="x", role="directors", active=True)
        product = Product(
            product_code="REC100",
            product_name="Reconciliation Product",
            hsn="0910",
            gst_rate=5,
            unit="Pcs",
            default_rate=100,
            tally_stock_item_name="Reconciliation Product",
        )
        db.add_all([admin, director, product])
        db.flush()
        first = Batch(
            batch_number="AUD-REC-001",
            batch_type=BatchType.AUDIT.value,
            user_id=admin.id,
            status=BatchStatus.SUBMITTED.value,
            submitted_at=datetime(2026, 6, 29, 10, 10, tzinfo=timezone.utc),
        )
        second = Batch(
            batch_number="AUD-REC-002",
            batch_type=BatchType.AUDIT.value,
            user_id=admin.id,
            status=BatchStatus.SUBMITTED.value,
            submitted_at=datetime(2026, 6, 29, 10, 40, tzinfo=timezone.utc),
        )
        outside = Batch(
            batch_number="AUD-REC-003",
            batch_type=BatchType.AUDIT.value,
            user_id=admin.id,
            status=BatchStatus.SUBMITTED.value,
            submitted_at=datetime(2026, 6, 29, 13, 0, tzinfo=timezone.utc),
        )
        db.add_all([first, second, outside])
        db.flush()
        db.add_all(
            [
                AuditFinding(
                    batch_id=first.id,
                    serial_number="REC100-000001",
                    product_code=product.product_code,
                    product_name=product.product_name,
                    finding_type="MISSING",
                    expected_status=SerialStatus.IN_STOCK.value,
                ),
                AuditFinding(
                    batch_id=second.id,
                    serial_number="REC100-000002",
                    product_code=product.product_code,
                    product_name=product.product_name,
                    finding_type="EXTRA",
                    expected_status=SerialStatus.IN_STOCK.value,
                    scanned_status=SerialStatus.SOLD.value,
                ),
                AuditFinding(
                    batch_id=outside.id,
                    serial_number="REC100-000003",
                    product_code=product.product_code,
                    product_name=product.product_name,
                    finding_type="MISSING",
                    expected_status=SerialStatus.IN_STOCK.value,
                ),
            ]
        )
        db.commit()

    def signed_request(user_id: int) -> Request:
        return Request(
            {
                "type": "http",
                "method": "GET",
                "path": "/reports/audit-reconciliation.xlsx",
                "headers": [(b"cookie", f"{SESSION_COOKIE}={create_session_token(user_id)}".encode())],
                "query_string": b"",
                "server": ("testserver", 80),
                "scheme": "http",
            }
        )

    with Session() as db:
        admin_response = audit_reconciliation_excel(
            signed_request(1),
            start="2026-06-29T10:00",
            end="2026-06-29T11:00",
            db=db,
        )
        director_response = audit_reconciliation_excel(
            signed_request(2),
            start="2026-06-29T10:00",
            end="2026-06-29T11:00",
            db=db,
        )
    engine.dispose()

    workbook = load_workbook(BytesIO(admin_response.body))
    summary = workbook["Summary"]
    products = workbook["Product Reconciliation"]

    assert admin_response.status_code == 200
    assert director_response.status_code == 200
    assert admin_response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert summary["B3"].value == 2
    assert summary["B5"].value == 1
    assert summary["B6"].value == 1
    assert products["A2"].value == "REC100"
    assert products["C2"].value == "AUD-REC-001, AUD-REC-002"
    assert products["E2"].value == 1
    assert products["F2"].value == 1


def test_loss_report_shows_factor_values_for_admin_and_super_admin():
    engine = create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    Session = sessionmaker(bind=engine)

    with Session() as db:
        admin = User(id=1, username="admin", password_hash="x", role="admin", active=True)
        root = User(id=2, username="root", password_hash="x", role="super_admin", active=True)
        sales = User(id=3, username="sales", password_hash="x", role="sales", active=True)
        product = Product(
            product_code="LOSS100",
            product_name="Loss test product",
            hsn="0910",
            gst_rate=5,
            unit="Pcs",
            default_rate=100,
            tally_stock_item_name="Loss test product",
        )
        db.add_all([admin, root, sales, product])
        db.flush()
        theft_serial = Serial(
            serial_number="LOSS100-000001",
            product_id=product.id,
            status=SerialStatus.ISSUED.value,
        )
        transport_serial = Serial(
            serial_number="LOSS100-000002",
            product_id=product.id,
            status=SerialStatus.ISSUED.value,
        )
        db.add_all([theft_serial, transport_serial])
        db.flush()
        theft_batch = Batch(
            batch_number="ISS-THEFT-001",
            batch_type=BatchType.ISSUE.value,
            reason_code="THEFT",
            user_id=admin.id,
            status=BatchStatus.SUBMITTED.value,
        )
        transport_batch = Batch(
            batch_number="ISS-TRANSPORT-001",
            batch_type=BatchType.ISSUE.value,
            reason_code="TRANSPORTATION",
            user_id=admin.id,
            status=BatchStatus.SUBMITTED.value,
        )
        db.add_all([theft_batch, transport_batch])
        db.flush()
        db.add_all(
            [
                BatchItem(batch_id=theft_batch.id, serial_id=theft_serial.id, quantity=1, rate=125.50),
                BatchItem(batch_id=transport_batch.id, serial_id=transport_serial.id, quantity=1),
                InventoryTransaction(
                    transaction_type=TransactionType.ISSUE.value,
                    serial_id=theft_serial.id,
                    product_id=product.id,
                    batch_id=theft_batch.id,
                    user_id=admin.id,
                    serial_number=theft_serial.serial_number,
                    status_from=SerialStatus.IN_STOCK.value,
                    status_to=SerialStatus.ISSUED.value,
                    reason_code="THEFT",
                ),
                InventoryTransaction(
                    transaction_type=TransactionType.ISSUE.value,
                    serial_id=transport_serial.id,
                    product_id=product.id,
                    batch_id=transport_batch.id,
                    user_id=admin.id,
                    serial_number=transport_serial.serial_number,
                    status_from=SerialStatus.IN_STOCK.value,
                    status_to=SerialStatus.ISSUED.value,
                    reason_code="TRANSPORTATION",
                ),
            ]
        )
        db.commit()
        save_role_access_config(db, {"reports_data": {"sales": "view"}})

    def override_get_db():
        db = Session()
        try:
            yield db
        finally:
            db.close()

    app.dependency_overrides[get_db] = override_get_db
    try:
        client = TestClient(app, follow_redirects=False, headers={"Origin": "http://testserver"})
        admin_response = client.get("/reports", cookies={SESSION_COOKIE: create_session_token(1)})
        root_response = client.get("/reports", cookies={SESSION_COOKIE: create_session_token(2)})
        sales_response = client.get("/reports", cookies={SESSION_COOKIE: create_session_token(3)})
    finally:
        app.dependency_overrides.clear()
        engine.dispose()

    for response in (admin_response, root_response):
        assert response.status_code == 200
        assert "<h2>Losses</h2>" in response.text
        assert "<th>Loss due to</th>" in response.text
        assert "Transportation" in response.text
        assert "Theft" in response.text
        assert "Other Things" in response.text
        assert "Rs 125.50" in response.text
        assert "Rs 225.50" in response.text
    assert sales_response.status_code == 200
    assert "<h2>Losses</h2>" not in sales_response.text


def test_dashboard_renders_stock_and_activity_charts():
    engine = create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    Session = sessionmaker(bind=engine)

    with Session() as db:
        user = User(id=1, username="admin", password_hash="x", role="admin", active=True)
        product = Product(
            product_code="SG200",
            product_name="Pepper",
            hsn="0910",
            gst_rate=5,
            unit="Pcs",
            default_rate=100,
            tally_stock_item_name="Pepper",
        )
        db.add_all([user, product])
        db.flush()
        serial = Serial(
            serial_number="SG200-000001",
            product_id=product.id,
            status=SerialStatus.IN_STOCK.value,
        )
        db.add(serial)
        db.flush()
        db.add(
            ScanLog(
                serial_id=serial.id,
                serial_number_raw=serial.serial_number,
                user_id=user.id,
                action=BatchType.AUDIT.value,
                status="SCANNED",
            )
        )
        db.commit()

    def override_get_db():
        db = Session()
        try:
            yield db
        finally:
            db.close()

    app.dependency_overrides[get_db] = override_get_db
    try:
        client = TestClient(app, follow_redirects=False, headers={"Origin": "http://testserver"})
        response = client.get("/", cookies={SESSION_COOKIE: create_session_token(1)})
        data_response = client.get("/dashboard/data", cookies={SESSION_COOKIE: create_session_token(1)})
    finally:
        app.dependency_overrides.clear()
        engine.dispose()

    assert response.status_code == 200
    assert "QR status mix" in response.text
    assert "Scan activity" in response.text
    assert "In Stock" in response.text
    assert data_response.status_code == 200
    assert "QR status mix" in data_response.json()["charts_html"]
