from datetime import date
from decimal import Decimal
from io import BytesIO

import pytest
from fastapi import HTTPException
from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from starlette.requests import Request

from app.auth import SESSION_COOKIE
from app.models import (
    BatchItem,
    BatchType,
    GstRegistrationType,
    GstTreatment,
    Product,
    ScanLog,
    SerialStatus,
    User,
)
from app.routers.batches import tally_excel_export
from app.security import create_session_token
from app.services.assignment import parse_bulk_assignment_xlsx
from app.services.inventory import InventoryError, add_serial_to_batch, create_batch, generate_serials
from app.services.sale_returns import scan_sale_return_product
from app.services.settings import update_settings
from app.services.tally_excel import (
    TALLY_ACCOUNTING_REQUIRED_EXPORT_FIELDS,
    TALLY_ACCOUNTING_VOUCHER_HEADERS,
    batch_tally_xlsx,
    import_tally_excel_to_batch,
    tally_accounting_default_deselected_fields,
)


VALID_TALLY_EXCEL_SETTINGS = {
    "sales_voucher_type": "Sales",
    "purchase_voucher_type": "Purchase",
    "sales_ledger_name": "Sales Ledger",
    "purchase_ledger_name": "Purchase Ledger",
    "cgst_ledger_name": "Input CGST @  2.5 %",
    "sgst_ledger_name": "Input SGST@2.5%",
    "sales_gst_ledger_mappings": "5 | Sales @ 5% | Output CGST @ 2.5% | Output SGST @ 2.5% | Output IGST @ 5%",
    "round_off_ledger_name": "ROUND OFF",
}


def _workbook_bytes(headers: list[str], rows: list[list[object]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _signed_request(user_id: int, batch_id: int) -> Request:
    return Request(
        {
            "type": "http",
            "method": "GET",
            "path": f"/batches/{batch_id}/tally.xlsx",
            "headers": [(b"cookie", f"{SESSION_COOKIE}={create_session_token(user_id)}".encode())],
            "query_string": b"",
            "server": ("testserver", 80),
            "scheme": "http",
        }
    )


def _product(code: str = "TALLYXL") -> Product:
    return Product(
        product_code=code,
        product_name="Tally Excel Product",
        hsn="0910",
        gst_rate=5,
        unit="Pcs",
        default_rate=100,
        tally_stock_item_name="Tally Excel Product",
    )


def test_tally_excel_default_fields_for_registered_interstate_sale(db_session):
    user = User(username="registered-interstate-xlsx", password_hash="x", role="sales")
    db_session.add(user)
    db_session.commit()
    batch = create_batch(
        db_session,
        user,
        BatchType.SALE,
        "Registered Buyer",
        "",
        party_state="Tamil Nadu",
        party_gst_registration_type=GstRegistrationType.REGULAR.value,
        party_gstin="33ABCDE1234F1Z5",
        gst_treatment=GstTreatment.INTER_STATE.value,
    )

    fields = tally_accounting_default_deselected_fields(batch)

    assert fields == ["CGST Rate", "SGST/UTGST Rate"]


def test_tally_excel_download_route_is_admin_only(db_session):
    admin = User(username="xlsx-admin", password_hash="x", role="admin", active=True)
    sales = User(username="xlsx-sales", password_hash="x", role="sales", active=True)
    product = _product("TALLY-ADMIN")
    db_session.add_all([admin, sales, product])
    db_session.commit()
    update_settings(db_session, VALID_TALLY_EXCEL_SETTINGS)
    serial = generate_serials(db_session, product, 1, initial_status=SerialStatus.IN_STOCK)[0]
    batch = create_batch(db_session, sales, BatchType.SALE, "Customer Ledger", "")
    add_serial_to_batch(db_session, batch, sales, serial.serial_number)

    admin_response = tally_excel_export(_signed_request(admin.id, batch.id), batch.id, db=db_session)

    assert admin_response.status_code == 200
    assert admin_response.body.startswith(b"PK")
    with pytest.raises(HTTPException) as exc:
        tally_excel_export(_signed_request(sales.id, batch.id), batch.id, db=db_session)
    assert exc.value.status_code == 403


def test_tally_excel_default_fields_for_registered_intrastate_sale(db_session):
    user = User(username="registered-intrastate-xlsx", password_hash="x", role="sales")
    db_session.add(user)
    db_session.commit()
    batch = create_batch(
        db_session,
        user,
        BatchType.SALE,
        "Registered Buyer",
        "",
        party_state="Karnataka",
        party_gst_registration_type=GstRegistrationType.REGULAR.value,
        party_gstin="29ABCDE1234F1Z5",
        gst_treatment=GstTreatment.INTRA_STATE.value,
    )

    fields = tally_accounting_default_deselected_fields(batch)

    assert fields == ["IGST Rate"]


def test_tally_excel_default_fields_for_unregistered_intrastate_sale(db_session):
    user = User(username="b2c-intrastate-xlsx", password_hash="x", role="sales")
    db_session.add(user)
    db_session.commit()
    batch = create_batch(
        db_session,
        user,
        BatchType.SALE,
        "Walk-in Buyer",
        "",
        party_state="Karnataka",
        party_gst_registration_type=GstRegistrationType.UNREGISTERED_CONSUMER.value,
        gst_treatment=GstTreatment.INTRA_STATE.value,
    )

    fields = tally_accounting_default_deselected_fields(batch)

    assert fields == [
        "Buyer/Supplier - Address",
        "Buyer/Supplier - Pincode",
        "Buyer/Supplier - GST Registration Type",
        "Buyer/Supplier - GSTIN/UIN",
        "HSN/SAC Details",
        "IGST Rate",
    ]
    assert not set(fields) & set(TALLY_ACCOUNTING_REQUIRED_EXPORT_FIELDS)


def test_tally_excel_default_fields_for_unregistered_interstate_sale(db_session):
    user = User(username="b2c-interstate-xlsx", password_hash="x", role="sales")
    db_session.add(user)
    db_session.commit()
    batch = create_batch(
        db_session,
        user,
        BatchType.SALE,
        "Walk-in Buyer",
        "",
        party_state="Tamil Nadu",
        party_gst_registration_type=GstRegistrationType.UNREGISTERED_CONSUMER.value,
        gst_treatment=GstTreatment.INTER_STATE.value,
    )

    fields = tally_accounting_default_deselected_fields(batch)

    assert fields == [
        "Buyer/Supplier - Address",
        "Buyer/Supplier - Pincode",
        "Buyer/Supplier - GST Registration Type",
        "Buyer/Supplier - GSTIN/UIN",
        "HSN/SAC Details",
        "CGST Rate",
        "SGST/UTGST Rate",
    ]


def test_tally_excel_default_fields_for_composition_sale(db_session):
    user = User(username="composition-xlsx", password_hash="x", role="sales")
    db_session.add(user)
    db_session.commit()
    batch = create_batch(
        db_session,
        user,
        BatchType.SALE,
        "Composition Buyer",
        "",
        party_state="Tamil Nadu",
        party_gst_registration_type=GstRegistrationType.COMPOSITION.value,
        party_gstin="33ABCDE1234F1Z5",
        gst_treatment=GstTreatment.INTER_STATE.value,
    )

    fields = tally_accounting_default_deselected_fields(batch)

    assert fields == [
        "GST Rate Details",
        "GST Taxability Type",
        "IGST Rate",
        "CGST Rate",
        "SGST/UTGST Rate",
    ]


def test_tally_excel_import_picks_fefo_stock_and_keeps_rate(db_session):
    user = User(username="admin", password_hash="x", role="admin")
    product = _product()
    db_session.add_all([user, product])
    db_session.commit()
    late = generate_serials(
        db_session,
        product,
        1,
        initial_status=SerialStatus.IN_STOCK,
        expiry_date=date(2026, 12, 31),
    )[0]
    early = generate_serials(
        db_session,
        product,
        1,
        initial_status=SerialStatus.IN_STOCK,
        expiry_date=date(2026, 8, 1),
    )[0]
    next_early = generate_serials(
        db_session,
        product,
        1,
        initial_status=SerialStatus.IN_STOCK,
        expiry_date=date(2026, 9, 1),
    )[0]
    batch = create_batch(db_session, user, BatchType.ISSUE, "Marketing", "", "SAMPLE")
    data = _workbook_bytes(
        ["Description of Goods", "Quantity", "Rate"],
        [[product.tally_stock_item_name, 2, 123.45]],
    )

    result = import_tally_excel_to_batch(db_session, batch, user, data)

    items = db_session.scalars(select(BatchItem).where(BatchItem.batch_id == batch.id)).all()
    logs = db_session.scalars(select(ScanLog).where(ScanLog.batch_id == batch.id)).all()
    assert result.quantity == 2
    assert {item.serial_id for item in items} == {early.id, next_early.id}
    assert late.id not in {item.serial_id for item in items}
    assert {item.rate for item in items} == {123.45}
    assert all(item.fefo_picked for item in items)
    assert {log.status for log in logs} == {"EXCEL_IMPORTED"}


def test_tally_excel_export_can_be_read_back_as_import_lines(db_session):
    user = User(username="admin", password_hash="x", role="admin")
    product = _product("TALLYX2")
    db_session.add_all([user, product])
    db_session.commit()
    generate_serials(db_session, product, 2, initial_status=SerialStatus.IN_STOCK)
    batch = create_batch(db_session, user, BatchType.ISSUE, "Marketing", "", "SAMPLE")
    import_tally_excel_to_batch(
        db_session,
        batch,
        user,
        _workbook_bytes(["Product Code", "Quantity", "Rate"], [[product.product_code, 2, 88.5]]),
    )

    data = batch_tally_xlsx(batch)
    workbook = load_workbook(BytesIO(data), data_only=True)
    sheet = workbook.active
    lines = parse_bulk_assignment_xlsx(db_session, data, allow_product_create=False)

    assert data.startswith(b"PK")
    assert sheet["B6"].value == "Description of Goods"
    assert sheet["B7"].value == product.tally_stock_item_name
    assert sheet["B2"].value == 1
    assert sheet["B4"].number_format == "DD-MM-YYYY"
    assert sheet["F7"].value == 2
    assert sheet["H7"].value == 88.5
    headers = [sheet.cell(6, c).value for c in range(1, sheet.max_column + 1)]
    amount_col = headers.index("Amount excl. GST") + 1
    assert sheet.cell(7, amount_col).value == 177
    assert len(lines) == 1
    assert lines[0].product.id == product.id
    assert lines[0].quantity == 2
    assert lines[0].rate == Decimal("88.5")


def test_sale_tally_excel_export_uses_tally_accounting_voucher_template(db_session):
    user = User(username="sales-xlsx", password_hash="x", role="sales")
    product = _product("TALLYSALE")
    db_session.add_all([user, product])
    db_session.commit()
    serials = generate_serials(db_session, product, 2, initial_status=SerialStatus.IN_STOCK)
    batch = create_batch(db_session, user, BatchType.SALE, "Customer Ledger", "")
    for serial in serials:
        add_serial_to_batch(db_session, batch, user, serial.serial_number)

    data = batch_tally_xlsx(batch, VALID_TALLY_EXCEL_SETTINGS)
    workbook = load_workbook(BytesIO(data), data_only=True)
    sheet = workbook.active
    headers = [sheet.cell(1, c).value for c in range(1, sheet.max_column + 1)]
    ledger_col = headers.index("Ledger Name") + 1
    amount_col = headers.index("Ledger Amount") + 1
    drcr_col = headers.index("Ledger Amount Dr/Cr") + 1
    item_col = headers.index("Item Name") + 1
    qty_col = headers.index("Billed Quantity") + 1
    rate_col = headers.index("Item Rate") + 1
    hsn_col = headers.index("HSN/SAC") + 1
    change_mode_col = headers.index("Change Mode ") + 1
    voucher_date_col = headers.index("Voucher Date") + 1
    voucher_number_col = headers.index("Voucher Number") + 1

    assert sheet.title == "Accounting Voucher"
    assert headers == TALLY_ACCOUNTING_VOUCHER_HEADERS
    assert "Description of Goods" not in headers
    assert "HSN Code" not in headers
    assert "GST Rate %" not in headers
    assert sheet.cell(2, ledger_col).value == "Customer Ledger"
    assert sheet.cell(2, voucher_date_col).number_format == "DD-MM-YYYY"
    assert sheet.cell(2, voucher_number_col).value == 1
    assert sheet.cell(2, amount_col).value == 200
    assert sheet.cell(2, drcr_col).value == "Dr"
    assert sheet.cell(3, ledger_col).value == "Sales @ 5%"
    assert sheet.cell(3, drcr_col).value == "Cr"
    assert sheet.cell(3, item_col).value == product.tally_stock_item_name
    assert sheet.cell(3, qty_col).value == 2
    assert sheet.cell(3, rate_col).value == 100
    assert sheet.cell(3, hsn_col).value == product.hsn
    assert sheet.cell(3, change_mode_col).value == "Accounting Invoice"
    assert sheet.cell(4, ledger_col).value == "Output CGST @ 2.5%"
    assert sheet.cell(5, ledger_col).value == "Output SGST @ 2.5%"


def test_unregistered_purchase_tally_excel_export_does_not_append_img_column(db_session):
    user = User(username="purchase-img-xlsx", password_hash="x", role="purchase")
    product = _product("TALLYIMG")
    db_session.add_all([user, product])
    db_session.commit()
    serial = generate_serials(db_session, product, 1)[0]
    batch = create_batch(
        db_session,
        user,
        BatchType.PURCHASE,
        "Farmer Supplier",
        "",
        party_gst_registration_type=GstRegistrationType.UNREGISTERED_CONSUMER.value,
    )
    add_serial_to_batch(db_session, batch, user, serial.serial_number)

    data = batch_tally_xlsx(batch, VALID_TALLY_EXCEL_SETTINGS)
    sheet = load_workbook(BytesIO(data), data_only=True).active
    headers = [sheet.cell(1, c).value for c in range(1, sheet.max_column + 1)]

    assert headers == TALLY_ACCOUNTING_VOUCHER_HEADERS
    assert sheet.cell(2, headers.index("Buyer/Supplier - GST Registration Type") + 1).value == "Unregistered/Consumer"


def test_registered_sale_tally_excel_export_does_not_append_img_column(db_session):
    user = User(username="registered-img-xlsx", password_hash="x", role="sales")
    product = _product("TALLYNOIMG")
    db_session.add_all([user, product])
    db_session.commit()
    serial = generate_serials(db_session, product, 1, initial_status=SerialStatus.IN_STOCK)[0]
    batch = create_batch(
        db_session,
        user,
        BatchType.SALE,
        "Registered Buyer",
        "",
        party_state="Karnataka",
        party_gst_registration_type=GstRegistrationType.REGULAR.value,
        party_gstin="29ABCDE1234F1Z5",
        gst_treatment=GstTreatment.INTRA_STATE.value,
    )
    add_serial_to_batch(db_session, batch, user, serial.serial_number)

    data = batch_tally_xlsx(batch, VALID_TALLY_EXCEL_SETTINGS)
    sheet = load_workbook(BytesIO(data), data_only=True).active
    headers = [sheet.cell(1, c).value for c in range(1, sheet.max_column + 1)]

    assert headers == TALLY_ACCOUNTING_VOUCHER_HEADERS


def test_tally_excel_export_can_include_selected_fields_only(db_session):
    user = User(username="selected-xlsx", password_hash="x", role="sales")
    product = _product("TALLYSELECT")
    db_session.add_all([user, product])
    db_session.commit()
    serial = generate_serials(db_session, product, 1, initial_status=SerialStatus.IN_STOCK)[0]
    batch = create_batch(db_session, user, BatchType.SALE, "Customer Ledger", "")
    add_serial_to_batch(db_session, batch, user, serial.serial_number)

    data = batch_tally_xlsx(
        batch,
        VALID_TALLY_EXCEL_SETTINGS,
        ["Voucher Type Name", "Voucher Number", "Ledger Name", "Item Name"],
        {
            "voucher_type": "Retail Sales",
            "voucher_number": "INV-SELECT-1",
            "party_ledger": "Walk-in Customer",
        },
    )
    sheet = load_workbook(BytesIO(data), data_only=True).active

    assert [cell.value for cell in sheet[1]] == [
        "Voucher Type Name",
        "Voucher Number",
        "Ledger Name",
        "Item Name",
    ]
    assert sheet["A2"].value == "Retail Sales"
    assert sheet["B2"].value == 1
    assert sheet["C2"].value == "Walk-in Customer"
    assert sheet["D3"].value == product.tally_stock_item_name


def test_sale_tally_excel_export_has_one_item_row_and_can_be_imported_by_app(db_session):
    user = User(username="sales-xlsx-roundtrip", password_hash="x", role="sales")
    product = _product("TALLYSALE2")
    db_session.add_all([user, product])
    db_session.commit()
    serials = generate_serials(db_session, product, 2, initial_status=SerialStatus.IN_STOCK)
    batch = create_batch(db_session, user, BatchType.SALE, "Customer Ledger", "")
    for serial in serials:
        add_serial_to_batch(db_session, batch, user, serial.serial_number)

    data = batch_tally_xlsx(batch, VALID_TALLY_EXCEL_SETTINGS)
    workbook = load_workbook(BytesIO(data), data_only=True)
    sheet = workbook.active
    headers = [sheet.cell(1, c).value for c in range(1, sheet.max_column + 1)]
    item_col = headers.index("Item Name") + 1
    qty_col = headers.index("Billed Quantity") + 1
    item_rows = [
        row
        for row in range(2, sheet.max_row + 1)
        if sheet.cell(row, item_col).value == product.tally_stock_item_name
    ]
    lines = parse_bulk_assignment_xlsx(db_session, data, allow_product_create=False)

    assert item_rows == [3]
    assert sheet.cell(item_rows[0], qty_col).value == 2
    assert len(lines) == 1
    assert lines[0].product.id == product.id
    assert lines[0].quantity == 2
    assert lines[0].rate == Decimal("100")


def test_tally_excel_import_rejects_purchase_batches(db_session):
    user = User(username="admin", password_hash="x", role="admin")
    product = _product("TALLYX3")
    db_session.add_all([user, product])
    db_session.commit()
    batch = create_batch(db_session, user, BatchType.PURCHASE, "Supplier", "")
    data = _workbook_bytes(["Product Code", "Quantity"], [[product.product_code, 1]])

    with pytest.raises(InventoryError, match="sale, issue, and purchase return"):
        import_tally_excel_to_batch(db_session, batch, user, data)


def test_sale_tally_excel_import_waits_for_pending_return_shelf_scan(db_session):
    user = User(username="sales", password_hash="x", role="sales")
    product = _product("TALLYX4")
    db_session.add_all([user, product])
    db_session.commit()
    serials = generate_serials(db_session, product, 2, initial_status=SerialStatus.IN_STOCK)
    batch = create_batch(db_session, user, BatchType.SALE, "Customer Ledger", "")
    add_serial_to_batch(db_session, batch, user, serials[0].serial_number)
    scan_sale_return_product(db_session, batch, user, serials[0].serial_number)
    data = _workbook_bytes(["Product Code", "Quantity"], [[product.product_code, 1]])

    with pytest.raises(InventoryError, match="shelf QR"):
        import_tally_excel_to_batch(db_session, batch, user, data)

    items = db_session.scalars(select(BatchItem).where(BatchItem.batch_id == batch.id)).all()
    assert items == []
